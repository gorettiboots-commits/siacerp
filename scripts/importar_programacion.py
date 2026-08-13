"""Importa 'Programación semanal.xlsx' a la base del SIAC.

El folio que aparece en el Excel es el folio de PROGRAMACIÓN (se asigna al
programar) y es distinto al folio de pedido (PED-XXXX). Se guarda en
programacion_lineas.folio_prog.

Estructura detectada por hoja (varía entre semanas):
  - Hojas antiguas: CLIENTE, FOLIO, MODELO, PARES, tallas, FECHA, TUBO, CHINELA
  - Hoja nueva (03-07 agosto 2026): CLIENTE, MODELO, PIEL, COLOR, FOLIO,
    tallas, TOTAL, FECHA

La reimportación borra y recarga programación (--force para confirmar).
"""
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import re

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.database.db_manager import DatabaseManager  # noqa: E402

XLSX = ROOT / "Programación semanal.xlsx"

_MESES = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}


def _fecha_inicio(nombre: str) -> str:
    """Deriva la fecha de inicio (YYYY-MM-DD) del nombre de la hoja.

    Acepta nombres como '29 Sep - 04 Oct' (sin año) o '05-10 Enero 2026'.
    Para las hojas sin año se asume el inicio del calendario actual de la
    programación: sep-dic -> año base, ene-ago -> año base + 1.
    """
    partes = re.split(r"[\s\-]+", nombre.strip())
    numeros = []
    mes = None
    for p in partes:
        k = p.lower()
        if k[:3] in _MESES:
            if mes is None:
                mes = _MESES[k[:3]]
            continue
        if re.fullmatch(r"\d{4}", k):
            numeros.append((p, "anio"))
        elif re.fullmatch(r"\d{1,2}", k):
            numeros.append((p, "dia"))
    dias = [int(n) for n, t in numeros if t == "dia"]
    anios = [int(n) for n, t in numeros if t == "anio"]
    if not dias or not mes:
        return ""
    dia = dias[0]
    if anios:
        anio = anios[0]
    elif mes >= 9:
        anio = 2025
    else:
        anio = 2026
    return f"{anio:04d}-{mes:02d}-{dia:02d}"


def _num(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


def _fecha(v) -> str:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return ""


def _texto(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _detectar_columnas(ws) -> dict:
    hdr = {c: ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)}
    cols = {
        "cliente": None, "folio": None, "modelo": None, "piel": None,
        "color": None, "pares": None, "total": None, "tubo": None,
        "chinela": None, "fecha": None,
    }
    for c, v in hdr.items():
        if v is None:
            continue
        key = str(v).strip().upper()
        if key == "CLIENTE":
            cols["cliente"] = c
        elif key == "FOLIO":
            cols["folio"] = c
        elif key == "MODELO":
            cols["modelo"] = c
        elif key == "PIEL":
            cols["piel"] = c
        elif key == "COLOR":
            cols["color"] = c
        elif key == "PARES":
            cols["pares"] = c
        elif key == "TOTAL":
            cols["total"] = c
        elif key == "TUBO":
            cols["tubo"] = c
        elif key == "CHINELA":
            cols["chinela"] = c
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v and "FECHA" in str(v).upper():
            cols["fecha"] = c
            break
    tallas = {}
    for c, v in hdr.items():
        if isinstance(v, bool):
            continue
        if isinstance(v, (int, float)) and str(v) not in ("",):
            tallas[c] = str(v)
    return cols, tallas


def main():
    db = DatabaseManager()
    n_existentes = db.fetch_one(
        "SELECT COUNT(*) AS n FROM programacion_semana")["n"]
    if n_existentes and "--force" not in sys.argv:
        print(f"ABORTADO: ya hay {n_existentes} semanas en la BD.")
        print("Ejecute con --force para borrar y recargar la programación "
              "(se perderán los estatus cambiados en la app).")
        sys.exit(1)

    if not XLSX.exists():
        print(f"No se encontró: {XLSX}")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)

    db.execute("DELETE FROM programacion_linea_tallas")
    db.execute("DELETE FROM programacion_lineas")
    db.execute("DELETE FROM programacion_semana")

    total_lineas = 0
    total_pares = 0
    omitidas = []

    for orden, nombre in enumerate(wb.sheetnames):
        ws = wb[nombre]
        cols, tallas = _detectar_columnas(ws)
        if cols["cliente"] is None:
            print(f"  - hoja '{nombre}': sin columna CLIENTE, omitida")
            continue
        cur = db.execute(
            "INSERT INTO programacion_semana (nombre, fecha_inicio, orden) "
            "VALUES (?, ?, ?)",
            (nombre, _fecha_inicio(nombre), orden))
        semana_id = cur.lastrowid

        n_lineas = 0
        for r in range(3, ws.max_row + 1):
            cliente = _texto(ws.cell(row=r, column=cols["cliente"]).value)
            if not cliente:
                continue
            modelo = _texto(ws.cell(row=r, column=cols["modelo"]).value) if cols["modelo"] else ""
            folio = _texto(ws.cell(row=r, column=cols["folio"]).value) if cols["folio"] else ""
            piel = _texto(ws.cell(row=r, column=cols["piel"]).value) if cols["piel"] else ""
            color = _texto(ws.cell(row=r, column=cols["color"]).value) if cols["color"] else ""
            fecha = _fecha(ws.cell(row=r, column=cols["fecha"]).value) if cols["fecha"] else ""
            tubo = _texto(ws.cell(row=r, column=cols["tubo"]).value) if cols["tubo"] else ""
            chinela = _texto(ws.cell(row=r, column=cols["chinela"]).value) if cols["chinela"] else ""

            pares_tallas = []
            for c, talla in tallas.items():
                n = _num(ws.cell(row=r, column=c).value)
                if n > 0:
                    pares_tallas.append((talla, n))

            total = None
            if cols["pares"]:
                total = _num(ws.cell(row=r, column=cols["pares"]).value)
            if cols["total"] and total in (None, 0):
                total = _num(ws.cell(row=r, column=cols["total"]).value)
            if not total:
                total = sum(n for _, n in pares_tallas)

            if not modelo and not folio and not pares_tallas:
                omitidas.append(f"{nombre}: fila {r} sin datos (cliente suelto) {cliente}")
                continue

            cur = db.execute(
                """INSERT INTO programacion_lineas
                   (semana_id, orden, folio_prog, cliente, modelo, piel, color,
                    fecha_prog, tubo, chinela, total_pares)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (semana_id, r - 2, folio, cliente, modelo, piel, color,
                 fecha, tubo, chinela, total))
            linea_id = cur.lastrowid
            for talla, n in pares_tallas:
                db.execute(
                    """INSERT INTO programacion_linea_tallas (linea_id, talla, orden, pares)
                       VALUES (?, ?, ?, ?)""",
                    (linea_id, talla, float(talla), n))
            n_lineas += 1
            total_lineas += 1
            total_pares += total
        print(f"  - {nombre}: {n_lineas} líneas")

    n_sem = db.fetch_one("SELECT COUNT(*) AS n FROM programacion_semana")["n"]
    n_lin = db.fetch_one("SELECT COUNT(*) AS n FROM programacion_lineas")["n"]
    n_tal = db.fetch_one("SELECT COUNT(*) AS n FROM programacion_linea_tallas")["n"]
    suma = db.fetch_one(
        "SELECT COALESCE(SUM(total_pares),0) AS n FROM programacion_lineas")["n"]

    print("\n===== RESULTADO DE IMPORTACIÓN =====")
    print(f"Semanas:       {n_sem} ({len(wb.sheetnames)} hojas)")
    print(f"Líneas:        {n_lin} (suma de hojas {total_lineas})")
    print(f"Renglones de talla: {n_tal}")
    print(f"Pares:         {suma} (validado {total_pares})")
    print(f"Filas omitidas: {len(omitidas)}")
    for o in omitidas[:30]:
        print(f"   - {o}")
    if len(omitidas) > 30:
        print(f"   ... y {len(omitidas) - 30} más")


if __name__ == "__main__":
    main()

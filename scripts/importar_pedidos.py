"""Importa PEDIDOS_ESKIN_NORMALIZADO.xlsx (hoja PEDIDOS) a la base del SIAC.

Crea clientes (hoja CLIENTES) y pedidos agrupados por (cliente, fecha de pedido).
Usa ClientesController / modelos para mantener el mismo flujo que la UI.

Filas omitidas (reportadas):
  - sin modelo y sin pares (artefactos/nombres)
  - sin modelo con pares pero sin fecha de pedido (subtotales de sección)

Filas sin modelo pero con fecha y pares se importan como modelo 'SIN MODELO'.
"""
import sys
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path

import openpyxl
import re

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from src.controllers.clientes_controller import ClientesController  # noqa: E402
from src.database.db_manager import DatabaseManager  # noqa: E402

XLSX = ROOT / "PEDIDOS_ESKIN_NORMALIZADO.xlsx"
OBS_RUIDO = {"PROGRAMADO", "Programado", ""}


def _fecha(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str) and v.strip():
        return v.strip()
    return ""


def _num(v):
    if v is None or v == "":
        return 0
    try:
        f = float(v)
        return int(round(f))
    except (TypeError, ValueError):
        return 0


def _clasificar(modelo, pares, fecha):
    if modelo:
        if pares > 0:
            return "importar"
        return "omitir_con_modelo_sin_pares"
    if pares == 0:
        return "omitir_sin_datos"
    if not fecha:
        return "omitir_subtotal"
    return "sin_modelo"


def main():
    db = DatabaseManager()
    n_cli = db.fetch_one("SELECT COUNT(*) AS n FROM clientes")["n"]
    n_ped = db.fetch_one("SELECT COUNT(*) AS n FROM pedidos_cliente")["n"]
    if n_cli or n_ped:
        print(f"ABORTADO: ya existen {n_cli} clientes y {n_ped} pedidos en la BD.")
        print("El importador solo corre sobre una BD vacía (clientes/pedidos).")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # ---- Catálogo de puntos: agregar tallas que falten ----
    puntos = db.fetch_all("SELECT id, punto, orden FROM puntos_catalogo ORDER BY orden")
    mapa_punto_id = {r["punto"]: r["id"] for r in puntos}
    tallas_hoja = [h.value for h in wb["PEDIDOS"][1]
                   if isinstance(h.value, str) and re.fullmatch(r"\d+(\.\d+)?", h.value)]
    agregados = []
    for t in sorted(tallas_hoja, key=float):
        if t not in mapa_punto_id:
            orden = float(t) + 1
            cur = db.execute(
                "INSERT INTO puntos_catalogo (punto, orden) VALUES (?, ?)", (t, orden))
            mapa_punto_id[t] = cur.lastrowid
            agregados.append(t)
    if agregados:
        print(f"Puntos agregados al catálogo: {agregados}")

    # ---- Clientes (hoja CLIENTES) ----
    clientes_nombre_comercial = {}
    for r in wb["CLIENTES"].iter_rows(min_row=2, values_only=True):
        if r[0]:
            clientes_nombre_comercial[str(r[0]).strip()] = str(r[1]).strip() if r[1] else ""

    # ---- Hoja PEDIDOS ----
    ws = wb["PEDIDOS"]
    hdr = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdr)}
    tallas_idx = [i for i, h in enumerate(hdr)
                  if isinstance(h, str) and re.fullmatch(r"\d+(\.\d+)?", h)]

    controller = ClientesController()
    cliente_ids: dict[str, int] = {}
    grupos: "OrderedDict[tuple, list]" = OrderedDict()
    omitidas = []
    no_enteros = set()
    lineas = 0
    pares_importados = 0

    def cliente_id(nombre: str) -> int:
        nombre = nombre.strip()
        if nombre not in cliente_ids:
            exist = db.fetch_one("SELECT id FROM clientes WHERE nombre = ?", (nombre,))
            if exist:
                cliente_ids[nombre] = exist["id"]
            else:
                cliente_ids[nombre] = controller.crear_cliente(
                    nombre,
                    nombre_comercial=clientes_nombre_comercial.get(nombre, ""),
                )
        return cliente_ids[nombre]

    for r in ws.iter_rows(min_row=2, values_only=True):
        modelo = str(r[idx["MODELO"]]).strip() if r[idx["MODELO"]] else ""
        piel = str(r[idx["PIEL"]] or "").strip()
        color = str(r[idx["COLOR"]] or "").strip()
        cliente = str(r[idx["CLIENTE"]] or "").strip()
        fecha = _fecha(r[idx["FECHA PEDIDO"]])
        fecha_prog = _fecha(r[idx["FECHA PROG"]])
        obs = str(r[idx["OBSERVACIONES"]] or "").strip()
        status = str(r[idx["STATUS"]] or "").strip().upper()

        pares_por_punto = {mapa_punto_id[t]: _num(r[i]) for i in tallas_idx for t in [hdr[i]]}
        for i in tallas_idx:
            v = r[i]
            if isinstance(v, float) and v != int(v):
                no_enteros.add((hdr[i], v))
        pares = sum(pares_por_punto.values())

        cat = _clasificar(modelo, pares, fecha)
        if cat == "importar":
            cid = cliente_id(cliente)
            clave = (cid, fecha)
            grupos.setdefault(clave, []).append({
                "modelo": modelo, "piel": piel, "color": color,
                "obs": obs, "fecha_prog": fecha_prog, "programado": status == "PROGRAMADO",
                "puntos": [{"punto_id": p, "pares": n}
                           for p, n in pares_por_punto.items() if n > 0],
            })
            lineas += 1
            pares_importados += pares
        elif cat == "sin_modelo":
            cid = cliente_id(cliente)
            clave = (cid, fecha)
            grupos.setdefault(clave, []).append({
                "modelo": "SIN MODELO", "piel": piel, "color": color,
                "obs": obs, "fecha_prog": fecha_prog, "programado": status == "PROGRAMADO",
                "puntos": [{"punto_id": p, "pares": n}
                           for p, n in pares_por_punto.items() if n > 0],
            })
            lineas += 1
            pares_importados += pares
            omitidas.append(f"SIN MODELO -> importado como 'SIN MODELO': {cliente} "
                            f"fecha={fecha} modelo_orig={modelo or r[idx['MODELO']]!r} pares={pares}")
        else:
            omitidas.append(f"{cat}: {cliente} fecha={fecha or '-'} modelo={modelo or '-'} "
                            f"pares={pares} obs={obs!r}")

    # ---- Crear pedidos por grupo (cliente, fecha) ----
    n_pedidos = 0
    for (cid, fecha), lineas_grupo in grupos.items():
        folio = controller.siguiente_folio()
        fecha_prog = next((l["fecha_prog"] for l in lineas_grupo if l["fecha_prog"]), "")
        programado = any(l["programado"] for l in lineas_grupo)
        obs_set = []
        for l in lineas_grupo:
            o = l["obs"]
            if o and o not in OBS_RUIDO and o not in obs_set:
                obs_set.append(o)
        controller.crear_pedido(
            folio=folio,
            cliente_id=cid,
            fecha_pedido=fecha,
            fecha_programado=fecha_prog,
            estatus="programado" if programado else "pendiente",
            observaciones="; ".join(obs_set),
            detalle=lineas_grupo,
        )
        n_pedidos += 1

    # ---- Validación ----
    total_db = db.fetch_one(
        "SELECT COALESCE(SUM(total_pares),0) AS n FROM pedidos_cliente")["n"]
    n_db_clientes = db.fetch_one("SELECT COUNT(*) AS n FROM clientes")["n"]
    n_db_lineas = db.fetch_one(
        "SELECT COUNT(*) AS n FROM detalle_pedido_cliente")["n"]

    print("\n===== RESULTADO DE IMPORTACIÓN =====")
    print(f"Clientes creados:            {len(cliente_ids)} (BD: {n_db_clientes})")
    print(f"Pedidos creados:             {n_pedidos}")
    print(f"Líneas de detalle:           {lineas} (BD: {n_db_lineas})")
    print(f"Pares importados (tallas):   {pares_importados} (BD total: {total_db})")
    if no_enteros:
        print(f"Aviso: pares con decimales redondeados: {sorted(no_enteros, key=str)}")
    print(f"Filas omitidas:              {len(omitidas)}")
    for o in omitidas:
        print(f"   - {o}")

    # Cruce con RESUMEN
    resumen = {}
    for r in wb["RESUMEN"].iter_rows(min_row=2, values_only=True):
        if r[0]:
            resumen[r[0]] = r[1]
    print("\n----- Contraste con RESUMEN -----")
    print(f"RESUMEN clientes únicos:  {resumen.get('Clientes únicos')}")
    print(f"RESUMEN pares suma tallas: {resumen.get('Pares (suma columnas talla)')} "
          f"(importado {pares_importados}, dif {-pares_importados + int(resumen.get('Pares (suma columnas talla)') or 0)})")


if __name__ == "__main__":
    main()

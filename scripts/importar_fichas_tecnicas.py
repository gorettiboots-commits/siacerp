"""Importador de fichas técnicas (kardex) desde archivos GORETTI*.xlsx.

Cada hoja del Excel es un modelo/variante (p. ej. GBC-01, GVM-01) con su
boleto técnico: encabezado (marca, talla, género, horma, moldura, construcción,
corrida, tacón...), secciones de materiales (CORTE, PESPUNTE & PRELIMINARES,
MONTADO/AVÍOS & ACABADO), notas y hasta 6 imágenes (que se guardan en la base
de datos como BLOB, el primer JPEG como imagen principal de la ficha).

Mapeo: la hoja se enlaza al `modelos` que coincida por `codigo`
(ESTILO SISTEMA) o por `nombre` (ESTILO MUESTRA). Si no existe, se CREA el
modelo. Es idempotente: al re-importar se reemplaza la ficha del modelo.

Uso:
    python scripts/importar_fichas_tecnicas.py [archivo.xlsx ...]
    python scripts/importar_fichas_tecnicas.py --dry-run
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from src.models.ficha_tecnica_model import FichaTecnicaModel
from src.models.produccion_model import ModeloModel

SECCIONES_CLAVE = {
    "CORTE": "CORTE",
    "PESPUNTE": "PESPUNTE & PRELIMINARES",
    "MONTADO": "MONTADO, AVÍOS & ACABADO",
    "NOTAS": "NOTAS Y REFERENCIAS",
}

CAMPOS_ENCABEZADO = {
    "ESTILO SISTEMA": "estilo_sistema",
    "ESTILO MUESTRA": "estilo_muestra",
    "MARCA": "marca",
    "TALLA": "talla",
    "GENERO": "genero",
    "HORMA": "horma",
    "MOLDURA": "moldura",
    "CONSTRUCCI": "construccion",
    "CORRIDA": "corrida",
    "SCALLOP": "scallop",
    "TAC": "tacon",  # TACÓN
}


def _txt(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor).replace("\u00c3\u0083", "").strip()
    return texto


def _encontrar_secciones(ws) -> list[dict]:
    """Localiza los títulos de sección por su texto en la columna B."""
    secciones = []
    for r in range(1, ws.max_row + 1):
        b = _txt(ws.cell(row=r, column=2).value).upper()
        for clave in SECCIONES_CLAVE:
            if b.startswith(clave):
                secciones.append({"fila": r, "clave": clave})
                break
    secciones.sort(key=lambda s: s["fila"])
    return secciones


def _extraer_encabezado(ws) -> dict:
    datos = {}
    for r in range(1, min(ws.max_row, 30) + 1):
        etiqueta = _txt(ws.cell(row=r, column=5).value).upper()
        valor = ws.cell(row=r, column=6).value
        for clave, campo in CAMPOS_ENCABEZADO.items():
            if clave in etiqueta and campo not in datos:
                datos[campo] = _txt(valor)
                break
    return datos


def _extraer_secciones(ws, secciones: list[dict]) -> list[dict]:
    """Convierte los bloques detectados en secciones con su detalle."""
    resultado = []
    notas = []
    for i, seccion in enumerate(secciones):
        clave = seccion["clave"]
        fila = seccion["fila"]
        siguiente_fila = (
            secciones[i + 1]["fila"] if i + 1 < len(secciones) else ws.max_row + 1
        )
        if clave == "NOTAS":
            for r in range(fila, min(fila + 4, siguiente_fila)):
                for c in (2, 4, 6):
                    v = _txt(ws.cell(row=r, column=c).value)
                    if v:
                        notas.append(v)
            continue

        nombre = SECCIONES_CLAVE[clave]
        detalle = []
        for r in range(fila + 1, siguiente_fila):
            if clave == "CORTE":
                c = _txt(ws.cell(row=r, column=3).value)
                d = _txt(ws.cell(row=r, column=4).value)
                prov = _txt(ws.cell(row=r, column=5).value)
                com = _txt(ws.cell(row=r, column=6).value)
                if c or d or prov or com:
                    detalle.append(
                        {"componente": c, "descripcion": d,
                         "proveedor": prov, "comentarios": com}
                    )
            else:
                c = _txt(ws.cell(row=r, column=3).value)
                d = _txt(ws.cell(row=r, column=4).value)
                if c or d:
                    detalle.append(
                        {"componente": c, "descripcion": d,
                         "proveedor": "", "comentarios": ""}
                    )
                c2 = _txt(ws.cell(row=r, column=6).value)
                d2 = _txt(ws.cell(row=r, column=7).value)
                if c2 or d2:
                    detalle.append(
                        {"componente": c2, "descripcion": d2,
                         "proveedor": "", "comentarios": ""}
                    )
        resultado.append({"nombre": nombre, "detalle": detalle})
    return resultado, " | ".join(notas)


def _extraer_imagen(ws) -> bytes | None:
    imagenes = getattr(ws, "_images", None) or []
    if not imagenes:
        return None
    datos = imagenes[0]._data()
    if isinstance(datos, bytes) and datos:
        return datos
    return None


def _crear_o_encontrar_modelo(m_modelo: ModeloModel, encabezado: dict,
                              codigo_preferido: str) -> int | None:
    nombre = _txt(encabezado.get("estilo_muestra"))
    # Se usa el nombre de hoja como `codigo` (identificador único del boleto)
    # porque en los archivos fuente el campo ESTILO SISTEMA puede estar
    # duplicado (p. ej. GBC-07/GBC-08 traen 'GBC-05'). El nombre de hoja es
    # siempre único por ficha.
    codigo = codigo_preferido
    if not codigo and not nombre:
        return None

    modelo_id = None
    if codigo:
        for m in m_modelo.listar():
            if _txt(m.get("codigo")).lower() == codigo.lower():
                modelo_id = m["id"]
                break
    # El nombre (ESTILO MUESTRA) se usa solo como respaldo cuando no hay
    # código de hoja; nunca debe absorber varios boletos en un mismo modelo.
    if modelo_id is None and not codigo and nombre:
        for m in m_modelo.listar():
            if _txt(m.get("nombre")).lower() == nombre.lower():
                modelo_id = m["id"]
                break
    if modelo_id is None:
        genero = _txt(encabezado.get("genero"))
        moldura = _txt(encabezado.get("moldura"))
        descripcion = " ".join(x for x in (genero, moldura) if x)
        modelo_id = m_modelo.crear(
            codigo=codigo or nombre, nombre=nombre or codigo,
            descripcion=descripcion,
        )
    return modelo_id


def importar_archivo(ruta: Path, solo_secos: bool, m_modelo: ModeloModel,
                     f_modelo: FichaTecnicaModel, resumen: dict) -> None:
    wb = openpyxl.load_workbook(str(ruta), data_only=True)
    try:
        for hoja in wb.sheetnames:
            if hoja.strip().lower().startswith("hoja"):
                continue
            ws = wb[hoja]
            encabezado = _extraer_encabezado(ws)
            secciones_brutas = _encontrar_secciones(ws)
            secciones, notas = _extraer_secciones(ws, secciones_brutas)
            imagen = _extraer_imagen(ws)
            codigo = _txt(hoja).strip() or _txt(encabezado.get("estilo_sistema"))

            if solo_secos:
                secciones_limpias = [
                    s for s in secciones if s["detalle"]
                ]
                resumen["hojas"] += 1
                print(
                    f"  [DRY] {codigo}: {_txt(encabezado.get('estilo_muestra'))}"
                    f" | secciones {len(secciones_limpias)}"
                    f" | filas {sum(len(s['detalle']) for s in secciones_limpias)}"
                    f" | imagen {'SÍ' if imagen else 'no'}"
                )
                continue

            modelo_id = _crear_o_encontrar_modelo(m_modelo, encabezado, codigo)
            if modelo_id is None:
                continue
            f_modelo.eliminar_por_modelo(modelo_id)
            f_modelo.guardar(
                modelo_id=modelo_id,
                datos={**encabezado, "notas": notas},
                secciones=[s for s in secciones if s["detalle"]],
                imagen=imagen,
                fuente_archivo=ruta.name,
            )
            resumen["fichas"] += 1
            print(
                f"  OK {codigo}: {_txt(encabezado.get('estilo_muestra'))}"
                f" -> modelo {modelo_id}"
            )
    finally:
        wb.close()


def principal(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Importa fichas técnicas GORETTI")
    parser.add_argument("archivos", nargs="*")
    parser.add_argument("--dry-run", action="store_true",
                        help="Solo muestra qué se importaría, sin escribir.")
    args = parser.parse_args(argv)

    archivos = [Path(a) for a in args.archivos]
    if not archivos:
        archivos = [
            Path("GORETTI BOTIN CHIMU.xlsx"),
            Path("GORETTI CASUAL TUTU.xlsx"),
            Path("GORRETTI VAQUERA CHIMU.xlsx"),
        ]
    archivos = [a for a in archivos if a.exists()]
    if not archivos:
        print("No se encontraron archivos GORETTI*.xlsx.")
        return 1

    m_modelo = ModeloModel()
    f_modelo = FichaTecnicaModel()
    resumen = {"fichas": 0, "hojas": 0}

    print(f"Cargando {len(archivos)} archivo(s)...")
    for ruta in archivos:
        print(f"\n{'-'*40}\nARCHIVO: {ruta.name}")
        importar_archivo(ruta, args.dry_run, m_modelo, f_modelo, resumen)

    print(f"\nResumen: {resumen['fichas']} ficha(s) importada(s), "
          f"{resumen['hojas']} hoja(s) en dry-run.")
    return 0


if __name__ == "__main__":
    sys.exit(principal(sys.argv[1:]))

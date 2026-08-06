# -*- coding: utf-8 -*-
"""Normaliza PEDIDOS ESKIN 2026.xlsx para validar antes de cargarlo al ERP.

Detecta el layout de cada hoja (3 formatos detectados), normaliza cliente,
modelo, piel y color (preservando acentos), descarta filas basura y produce
PEDIDOS_ESKIN_NORMALIZADO.xlsx con las hojas:
RESUMEN, CLIENTES, MODELOS, PIELES, COLORES, PEDIDOS (detalle) y REVISAR.
"""

import sys
import unicodedata
from collections import Counter
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

RAIZ = Path(__file__).resolve().parent.parent
ENTRADA = RAIZ / "PEDIDOS ESKIN 2026.xlsx"
SALIDA = RAIZ / "PEDIDOS_ESKIN_NORMALIZADO.xlsx"

SUFIJOS_HOJA = (" 12-14", " 22-25", " 22-26", " 22 - 25", " 22-26 12-14")

# Correcciones de piel, con clave SIN acentos
PIEL_CORRECCIONES = {
    "PTON": "PITÓN",
    "PITON": "PITÓN",
    "FLOTER": "FLOTHER",
    "PULL - UP": "PULL UP",
}

# Correcciones de color, con clave SIN acentos
COLOR_CORRECCIONES = {
    "CAAMEL/CAMEL": "CAMEL/CAMEL",
    "CAME/CAMEL": "CAMEL/CAMEL",
    "CONAZ/AZUL": "COÑAC/AZUL",
    "NEGRO/NGO": "NEGRO/NEGRO",
    "BYN/NGO": "BYN/NEGRO",
    "CAFE-/CAFE": "CAFÉ/CAFÉ",
}

CAB_MODELO = ("MODELO",)
CAB_PIEL = ("PIEL", "MATERIAL")
CAB_COLOR = ("COLOR",)
CAB_FECHA_PEDIDO = ("FECHA DE PEDIDO", "FECHA")
CAB_FECHA_PROG = ("FECHA PROGRAMADO", "FECHA PROGRA", "FECHA PROGR", "FECHA P",
                  "FECHA PROG.", "FECHA DE PRO", "FECHA PROG-")
CAB_OBS = ("OBSERVACIONES", "OBSEVACION", "OBERVACION")
CAB_CLIENTE = ("CLIENTE",)
CAB_MARCA = ("MARCA",)
CAB_ROTULAR = ("ROTULAR",)
CAB_STATUS = ("STATUS",)
CAB_TOTAL = ("TOTAL",)


def sin_acentos(s):
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def norm_preserva(s):
    if s is None:
        return ""
    return " ".join(str(s).split())


def norm_alta(s):
    return sin_acentos(norm_preserva(s)).upper()


def nombre_cliente_de_hoja(hoja):
    n = norm_preserva(hoja)
    for suf in SUFIJOS_HOJA:
        if n.upper().endswith(suf.upper()):
            n = n[: -len(suf)].strip()
            break
    while n.startswith('"') and n.endswith('"') and len(n) >= 2:
        n = n[1:-1].strip()
    return n


def a_fecha(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def deteccion_columnas(hdr):
    pos = {k: None for k in ("obs", "cliente", "f_pedido", "f_prog", "modelo",
                             "piel", "color", "marca", "rotular", "total", "status")}
    tallas = []
    for j, h in enumerate(hdr, start=1):
        if h is None:
            continue
        if isinstance(h, (int, float)):
            tallas.append((h, j))
            continue
        t = norm_alta(h)
        if not t:
            continue
        if t in CAB_MODELO and pos["modelo"] is None:
            pos["modelo"] = j
        elif any(t.startswith(p) for p in CAB_PIEL) and pos["piel"] is None:
            pos["piel"] = j
        elif t in CAB_COLOR and pos["color"] is None:
            pos["color"] = j
        elif any(t.startswith(p) for p in CAB_OBS) and pos["obs"] is None:
            pos["obs"] = j
        elif t in CAB_CLIENTE and pos["cliente"] is None:
            pos["cliente"] = j
        elif t in CAB_MARCA and pos["marca"] is None:
            pos["marca"] = j
        elif t in CAB_ROTULAR and pos["rotular"] is None:
            pos["rotular"] = j
        elif t in CAB_TOTAL and pos["total"] is None:
            pos["total"] = j
        elif t in CAB_STATUS and pos["status"] is None:
            pos["status"] = j
        elif any(t.startswith(p) for p in CAB_FECHA_PEDIDO) and pos["f_pedido"] is None:
            pos["f_pedido"] = j
        elif any(t.startswith(p) for p in CAB_FECHA_PROG) and pos["f_prog"] is None:
            pos["f_prog"] = j
    tallas.sort(key=lambda x: x[0])
    return pos, [j for _, j in tallas], [v for v, _ in tallas]


def es_codigo_modelo(m):
    if m is None:
        return False
    if isinstance(m, (int, float)):
        return 100 <= m <= 99999
    base = norm_alta(m).split("-")[0].split(" ")[0]
    return base.isdigit() and 100 <= int(base) <= 99999


def normalizar_piel(v):
    raw = norm_preserva(v)
    if not raw:
        return "", ""
    key = norm_alta(raw)
    limpio = PIEL_CORRECCIONES.get(key, raw.upper())
    flags = []
    if key in PIEL_CORRECCIONES:
        flags.append("piel: tipográfico corregido")
    if key in ("MODELO", "PIEL", "COLOR", "CLIENTE"):
        flags.append("piel: es texto de cabecera")
    return limpio, " | ".join(flags)


def normalizar_color(v):
    raw = norm_preserva(v)
    if not raw:
        return "", ""
    raw = raw.replace("´", "").replace("`", "")
    partes = raw.split()
    if len(partes) == 2 and sin_acentos(partes[0]).upper() == sin_acentos(partes[1]).upper():
        raw = "/".join(partes)
    key = norm_alta(raw)
    limpio = COLOR_CORRECCIONES.get(key, raw.upper())
    flags = []
    if key in COLOR_CORRECCIONES:
        flags.append("color: tipográfico corregido")
    if key in ("MODELO", "PIEL", "COLOR", "CLIENTE", "TOTAL"):
        flags.append("color: es texto de cabecera")
    return limpio, " | ".join(flags)


def main() -> None:
    wb = openpyxl.load_workbook(ENTRADA, data_only=True)

    todas_tallas = {}
    for s in wb.sheetnames:
        ws = wb[s]
        hr = 1
        if ws.max_row >= 2:
            hdr2 = [ws.cell(2, j).value for j in range(1, ws.max_column + 1)]
            if any(norm_alta(h) in CAB_MODELO for h in hdr2 if h is not None):
                hr = 2
        hdr = [ws.cell(hr, j).value for j in range(1, ws.max_column + 1)]
        _, _, vals_tallas = deteccion_columnas(hdr)
        for v in vals_tallas:
            todas_tallas[v] = True
    lista_tallas = sorted(todas_tallas.keys())

    filas = []
    revisar = []
    for s in wb.sheetnames:
        ws = wb[s]
        hr = 1
        if ws.max_row >= 2:
            hdr2 = [ws.cell(2, j).value for j in range(1, ws.max_column + 1)]
            if any(norm_alta(h) in CAB_MODELO for h in hdr2 if h is not None):
                hr = 2
        hdr = [ws.cell(hr, j).value for j in range(1, ws.max_column + 1)]
        pos, cols_tallas, vals_tallas = deteccion_columnas(hdr)

        cliente = nombre_cliente_de_hoja(s)
        nombre_completo = ""
        if hr == 2:
            for j in range(1, ws.max_column + 1):
                v = ws.cell(1, j).value
                if isinstance(v, str) and norm_preserva(v):
                    txt = norm_preserva(v)
                    if norm_alta(txt) not in (norm_alta(cliente), "PEDIDO", "PEDIDOS"):
                        nombre_completo = txt
                        break

        for i in range(hr + 1, ws.max_row + 1):
            def celda(k):
                return ws.cell(i, pos[k]).value if pos.get(k) else None

            obs_raw = celda("obs")
            f_ped = a_fecha(celda("f_pedido"))
            f_prog = a_fecha(celda("f_prog"))
            modelo_raw = celda("modelo")
            piel_raw = celda("piel")
            color_raw = celda("color")
            cliente_row = norm_preserva(celda("cliente"))
            marca = norm_preserva(celda("marca"))
            rotular = norm_preserva(celda("rotular"))
            status = norm_alta(celda("status"))
            total_raw = celda("total")

            # cabecera repetida dentro de la hoja -> se descarta
            if norm_alta(modelo_raw) == "MODELO":
                continue

            pares_por_talla = {}
            suma_tallas = 0
            for t, j in zip(vals_tallas, cols_tallas):
                v = ws.cell(i, j).value
                if isinstance(v, (int, float)) and v:
                    pares_por_talla[t] = v
                    suma_tallas += v

            total = None
            if isinstance(total_raw, (int, float)):
                total = total_raw

            # fila vacía de sobra
            if modelo_raw is None and suma_tallas == 0 and total in (None, "", 0):
                continue

            flags = []
            modelo = ""
            if es_codigo_modelo(modelo_raw):
                if isinstance(modelo_raw, (int, float)):
                    modelo = str(int(modelo_raw))
                else:
                    modelo = norm_preserva(modelo_raw).split(" ")[0]
            elif isinstance(modelo_raw, (datetime, date)):
                flags.append("modelo: celda con fecha")
            elif modelo_raw is not None and norm_preserva(modelo_raw):
                flags.append("modelo: valor no es código (%s)"
                             % norm_preserva(modelo_raw)[:20])

            piel, f_piel = normalizar_piel(piel_raw)
            if f_piel:
                flags.append(f_piel)
            color, f_color = normalizar_color(color_raw)
            if f_color:
                flags.append(f_color)

            if total is not None and suma_tallas and abs(total - suma_tallas) > 0:
                flags.append("total difiere de la suma de tallas")
            if total is None and suma_tallas == 0:
                flags.append("sin tallas ni total")
            if not modelo and (suma_tallas or total):
                flags.append("sin modelo válido")

            fila = {
                "hoja": s,
                "cliente": cliente,
                "nombre_completo": nombre_completo,
                "cliente_row": cliente_row,
                "obs": norm_preserva(obs_raw),
                "f_pedido": f_ped,
                "f_prog": f_prog,
                "modelo": modelo,
                "modelo_raw": norm_preserva(modelo_raw),
                "piel": piel,
                "piel_raw": norm_preserva(piel_raw),
                "color": color,
                "color_raw": norm_preserva(color_raw),
                "marca": marca,
                "rotular": rotular,
                "status": status,
                "tallas": pares_por_talla,
                "suma_tallas": suma_tallas,
                "total": total,
                "flags": flags,
            }
            filas.append(fila)
            if flags:
                revisar.append(fila)

    n_lineas = len(filas)
    n_revisar = len(revisar)
    pares_total = sum(f["total"] for f in filas if f["total"] is not None)
    pares_suma = sum(f["suma_tallas"] for f in filas)
    clientes_n = len({f["cliente"] for f in filas})
    modelos_n = len({f["modelo"] for f in filas if f["modelo"]})
    piels_n = len({f["piel"] for f in filas if f["piel"]})
    colores_n = len({f["color"] for f in filas if f["color"]})

    por_cliente = Counter()
    pares_cliente = Counter()
    por_modelo = Counter()
    por_piel = Counter()
    por_color = Counter()
    for f in filas:
        por_cliente[f["cliente"]] += 1
        p = f["total"] if f["total"] is not None else f["suma_tallas"]
        pares_cliente[f["cliente"]] += p
        if f["modelo"]:
            por_modelo[f["modelo"]] += 1
        if f["piel"]:
            por_piel[f["piel"]] += 1
        if f["color"]:
            por_color[f["color"]] += 1

    out = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1e293b")
    hdr_font = Font(color="FFFFFF", bold=True)

    def estilizar(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(1, c)
            cell.fill = hdr_fill
            cell.font = hdr_font
        ws.freeze_panes = "A2"

    ws = out.active
    ws.title = "RESUMEN"
    for r, (k, v) in enumerate([
        ("Archivo origen", ENTRADA.name),
        ("Hojas", len(wb.sheetnames)),
        ("Líneas de pedido", n_lineas),
        ("Líneas para revisar", n_revisar),
        ("Pares (columna TOTAL)", round(pares_total, 1)),
        ("Pares (suma columnas talla)", round(pares_suma, 1)),
        ("Diferencia TOTAL vs suma tallas", round(pares_total - pares_suma, 1)),
        ("Clientes únicos", clientes_n),
        ("Modelos únicos", modelos_n),
        ("Pieles únicas", piels_n),
        ("Colores únicos", colores_n),
    ], start=1):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20

    ws = out.create_sheet("CLIENTES")
    ws.append(["CLIENTE", "NOMBRE/TÍTULO", "LINEAS", "PARES"])
    for c, n in por_cliente.most_common():
        nc = next((f["nombre_completo"] for f in filas
                   if f["cliente"] == c and f["nombre_completo"]), "")
        ws.append([c, nc, n, round(pares_cliente[c], 1)])
    estilizar(ws, 4)

    ws = out.create_sheet("MODELOS")
    ws.append(["MODELO", "LINEAS"])
    for m, n in por_modelo.most_common():
        ws.append([m, n])
    estilizar(ws, 2)

    ws = out.create_sheet("PIELES")
    ws.append(["PIEL", "LINEAS"])
    for p, n in por_piel.most_common():
        ws.append([p, n])
    estilizar(ws, 2)

    ws = out.create_sheet("COLORES")
    ws.append(["COLOR", "LINEAS"])
    for co, n in por_color.most_common():
        ws.append([co, n])
    estilizar(ws, 2)

    ws = out.create_sheet("PEDIDOS")
    cab = (["HOJA", "CLIENTE", "CLIENTE RENGLÓN", "OBSERVACIONES",
            "FECHA PEDIDO", "FECHA PROG", "MODELO", "PIEL", "COLOR",
            "MARCA", "ROTULAR", "STATUS"]
           + [str(t) for t in lista_tallas]
           + ["TOTAL", "SUMA TALLAS", "FLAGS"])
    ws.append(cab)
    for f in filas:
        row = [f["hoja"], f["cliente"], f["cliente_row"], f["obs"],
               f["f_pedido"], f["f_prog"], f["modelo"], f["piel"], f["color"],
               f["marca"], f["rotular"], f["status"]]
        row += [f["tallas"].get(t, "") for t in lista_tallas]
        row += [f["total"] if f["total"] is not None else "",
                f["suma_tallas"] or "",
                " | ".join(f["flags"])]
        ws.append(row)
    estilizar(ws, len(cab))

    ws = out.create_sheet("REVISAR")
    ws.append(cab)
    for f in revisar:
        row = [f["hoja"], f["cliente"], f["cliente_row"], f["obs"],
               f["f_pedido"], f["f_prog"], f["modelo"], f["piel"], f["color"],
               f["marca"], f["rotular"], f["status"]]
        row += [f["tallas"].get(t, "") for t in lista_tallas]
        row += [f["total"] if f["total"] is not None else "",
                f["suma_tallas"] or "",
                " | ".join(f["flags"])]
        ws.append(row)
    estilizar(ws, len(cab))

    out.save(SALIDA)

    print(f"Líneas: {n_lineas} | Revisar: {n_revisar} | Pares(TOTAL): {pares_total:,.1f} | "
          f"Pares(suma tallas): {pares_suma:,.1f}")
    print(f"Clientes: {clientes_n} | Modelos: {modelos_n} | Pieles: {piels_n} | Colores: {colores_n}")
    print("Salida:", SALIDA)


if __name__ == "__main__":
    main()

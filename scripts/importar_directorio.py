"""Importa proveedores e insumos desde DIRECTORIO.xlsx.

Hoja 1 (DIRECTOS):  proveedores (nombre, razon social, telefono, email, direccion).
Hoja 2 (MATERIALES): materiales por proveedor (material, unidad, precio, comentario).

El proceso es idempotente: si el proveedor o el insumo ya existe no lo duplica,
solo agrega lo que falte. Puede re-ejecutarse cada vez que se complete el Excel.
"""

import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.database.db_manager import DatabaseManager

EXCEL = Path(__file__).resolve().parent.parent / "DIRECTORIO.xlsx"
HOJA_PROVEEDORES = "DIRECTOS"
HOJA_MATERIALES = "MATERIALES "

STOP_RFC = {"SA", "S.A", "S.A.", "S.A.DE", "S.A. DE", "C.V", "C.V.", "CV",
            "DE", "LA", "DEL", "LOS", "LAS", "Y", "E", "EL", "A", "N/A", "NA"}

# Aliases de hoja2 -> nombre/razon social de hoja1 que no coinciden literalmente
ALIASES = {
    "PELETERIA ANELY": "PELETERIA ANELY",
    "SULTANA": "PIELES LA SULTANA",
    "PELETERIA EL TUKAN": "TUKAN",
    "JULIO HERRAJES Y ADORNOS": "JULIO HERRAJES Y ADORNO",
    "ETI-PLAST": "ETI-PLAST",
}

UNIDADES_EXCEL = {
    "PZA": "pieza",
    "PZAO": "pieza",
    "PAR": "par",
    "METRO": "metro",
    "MTRO": "metro",
    "M2": "dm2",
    "LAMINA": "lamina",
    "LATA": "lata",
}


def norm(s):
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = " ".join(s.strip().split())
    return s.replace(" )", ")").replace(" ,", ",")


def unidad_normalizada(texto):
    if not texto:
        return "pieza"
    limpio = "".join(c for c in norm(texto) if c.isalnum()).upper()
    return UNIDADES_EXCEL.get(limpio, limpio.lower())


def categoria_de_material(texto):
    t = norm(texto).upper()
    reglas = [
        ("SINTETICO", "SINTETICO"), ("CHAROL", "CHAROL"), ("CABRA", "CABRA"),
        ("SUELA", "SUELA"), ("CIERRE", "CIERRE"), ("MALLA", "MALLA"),
        ("CASCO", "CASCO"), ("CONTRAFUERTE", "CONTRAFUERTE"),
        ("APLICACIONES", "APLICACIONES"), ("HEBILLA", "HEBILLA"),
        ("PLANTA", "PLANTA"), ("HORMA", "HORMA"),
        ("ADHESIVOS Y QUIMICOS", "PEGAMENTO"),
        ("ADHESIVOS Y QUIMICOS", "ADHESIVO"),
        ("APLICACIONES", "PLASTISOL"), ("APLICACIONES", "TRANSFER"),
    ]
    for cat, key in reglas:
        if key in t:
            return cat
    return "MATERIA PRIMA"


def compact(s):
    return norm(s).replace(" ", "")


def generar_rfc(nombre, seq):
    palabras = [w for w in norm(nombre).split() if w not in STOP_RFC]
    letras = "".join(w[0] for w in palabras if w[0].isalnum()) if palabras else "RFC"
    letras = (letras + "XXX")[:3]
    return f"{letras}000101{seq:03d}"


def main() -> None:
    if not EXCEL.exists():
        print(f"No se encontró el archivo: {EXCEL}")
        return

    db = DatabaseManager()
    conn = db.connect()
    conn.execute("PRAGMA busy_timeout=8000")
    cursor = conn.cursor()

    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    # ---------- 1. Proveedores desde hoja DIRECTOS ----------
    ws1 = wb[HOJA_PROVEEDORES]
    proveedores = {}
    for i in range(3, ws1.max_row + 1):
        alias = ws1.cell(i, 3).value
        if not alias or not norm(alias):
            continue
        razon = ws1.cell(i, 4).value
        nombre = norm(razon) or norm(alias)
        if nombre in ("N/A", "NA", "X"):
            nombre = norm(alias)
        clave = norm(alias)
        if clave in proveedores:
            continue
        proveedores[clave] = {
            "nombre": nombre,
            "alias": norm(alias),
            "telefono": norm(ws1.cell(i, 7).value),
            "email": norm(ws1.cell(i, 8).value),
            "direccion": norm(ws1.cell(i, 9).value),
            "material_s1": norm(ws1.cell(i, 2).value),
        }
    print(f"Proveedores detectados en hoja 1: {len(proveedores)}")

    proveedor_ids = {}
    compact_ids = {}
    for clave, prov in proveedores.items():
        row = cursor.execute(
            "SELECT id FROM proveedores WHERE nombre IN (?, ?) LIMIT 1",
            (prov["nombre"], prov["alias"]),
        ).fetchone()
        if row:
            proveedor_ids[clave] = row[0]
            cursor.execute(
                "UPDATE proveedores SET telefono=?, email=?, direccion=? WHERE id=?",
                (prov["telefono"], prov["email"], prov["direccion"], row[0]),
            )
            compact_ids[compact(prov["alias"])] = row[0]
            compact_ids[compact(prov["nombre"])] = row[0]
            continue
        seq = 1
        rfc = generar_rfc(prov["nombre"], seq)
        while cursor.execute("SELECT 1 FROM proveedores WHERE rfc = ?", (rfc,)).fetchone():
            seq += 1
            rfc = generar_rfc(prov["nombre"], seq)
        cur = cursor.execute(
            "INSERT INTO proveedores (rfc, nombre, telefono, email, direccion) "
            "VALUES (?, ?, ?, ?, ?)",
            (rfc, prov["nombre"], prov["telefono"], prov["email"], prov["direccion"]),
        )
        proveedor_ids[clave] = cur.lastrowid
        compact_ids[compact(prov["alias"])] = cur.lastrowid
        compact_ids[compact(prov["nombre"])] = cur.lastrowid
        print(f"  + Proveedor: {prov['nombre']} (rfc {rfc})")

    def resolver_proveedor(alias_s2):
        a = norm(alias_s2)
        ca = compact(a)
        if ca in compact_ids:
            return compact_ids[ca]
        for ck, pid in compact_ids.items():
            if ca in ck or ck in ca:
                return pid
        target = ALIASES.get(a)
        if target:
            ct = compact(target)
            if ct in compact_ids:
                return compact_ids[ct]
            for ck, pid in compact_ids.items():
                if ct in ck or ck in ct:
                    return pid
        return None

    # ---------- 2. Insumos + proveedor_insumos desde hoja MATERIALES ----------
    ws2 = wb[HOJA_MATERIALES]
    next_codigo = 1
    filas = []
    for i in range(3, ws2.max_row + 1):
        alias = ws2.cell(i, 1).value
        material = ws2.cell(i, 2).value
        if not alias or not material or not norm(material):
            continue
        filas.append({
            "proveedor": resolver_proveedor(alias),
            "material": norm(material),
            "unidad": unidad_normalizada(ws2.cell(i, 3).value),
            "precio": ws2.cell(i, 4).value,
            "comentario": norm(ws2.cell(i, 6).value),
        })

    # registro de unidades de medida nuevas
    for f in filas:
        if not cursor.execute(
                "SELECT 1 FROM unidades_medida WHERE abreviatura = ?", (f["unidad"],)
        ).fetchone():
            cursor.execute(
                "INSERT INTO unidades_medida (nombre, abreviatura) VALUES (?, ?)",
                (f["unidad"].capitalize(), f["unidad"]),
            )

    # registro de insumos
    insumo_ids = {}
    por_nombre = {}
    for r in cursor.execute("SELECT id, nombre FROM insumos").fetchall():
        por_nombre[norm(r[1])] = r[0]
    for r in cursor.execute("SELECT codigo FROM insumos").fetchall():
        try:
            n = int(str(r[0]).split("IM")[-1])
            if r[0].startswith("IM") and n >= next_codigo:
                next_codigo = n + 1
        except (ValueError, IndexError):
            pass

    for f in filas:
        prov_clave = None
        for clave, pid in proveedor_ids.items():
            if pid == f["proveedor"]:
                prov_clave = clave
                break
        categoria = categoria_de_material(
            proveedores[prov_clave]["material_s1"]) if prov_clave else "MATERIA PRIMA"
        if f["material"] in por_nombre:
            iid = por_nombre[f["material"]]
            insumo_ids[f["material"]] = iid
            cur = cursor.execute(
                "SELECT codigo FROM insumos WHERE id = ?", (iid,)).fetchone()
            if cur and str(cur[0]).startswith("IM"):
                cursor.execute(
                    "UPDATE insumos SET categoria=?, unidad_medida=? WHERE id=?",
                    (categoria, f["unidad"], iid),
                )
            continue
        codigo = f"IM{next_codigo:03d}"
        while cursor.execute("SELECT 1 FROM insumos WHERE codigo = ?", (codigo,)).fetchone():
            next_codigo += 1
            codigo = f"IM{next_codigo:03d}"
        next_codigo += 1
        cur = cursor.execute(
            "INSERT INTO insumos (codigo, nombre, categoria, unidad_medida, stock_minimo) "
            "VALUES (?, ?, ?, ?, 0)",
            (codigo, f["material"], categoria, f["unidad"]),
        )
        insumo_ids[f["material"]] = cur.lastrowid
        por_nombre[f["material"]] = cur.lastrowid
        print(f"  + Insumo: {f['material']} [{categoria}] ({f['unidad']})")

    # ---------- 3. Relación proveedor <-> insumo ----------
    insertados = 0
    para_mensaje = []
    for f in filas:
        if f["proveedor"] is None:
            para_mensaje.append(f"  ! Sin proveedor resuelto: {f['material']}")
            continue
        insumo_id = insumo_ids.get(f["material"])
        if not insumo_id:
            continue
        ya = cursor.execute(
            "SELECT 1 FROM proveedor_insumos WHERE proveedor_id = ? AND insumo_id = ?",
            (f["proveedor"], insumo_id),
        ).fetchone()
        if ya:
            ins = cursor.execute(
                "SELECT codigo FROM insumos WHERE id = ?", (insumo_id,)).fetchone()
            if ins and str(ins[0]).startswith("IM"):
                precio = float(f["precio"]) if f["precio"] is not None else None
                cursor.execute(
                    "UPDATE proveedor_insumos SET unidad_medida=?, "
                    "precio=COALESCE(?, precio), comentario=? WHERE proveedor_id=? AND insumo_id=?",
                    (f["unidad"], precio, f["comentario"],
                     f["proveedor"], insumo_id),
                )
            continue
        precio = float(f["precio"]) if f["precio"] is not None else 0.0
        cursor.execute(
            "INSERT INTO proveedor_insumos "
            "(proveedor_id, insumo_id, unidad_medida, precio, comentario) "
            "VALUES (?, ?, ?, ?, ?)",
            (f["proveedor"], insumo_id, f["unidad"], precio, f["comentario"]),
        )
        insertados += 1

    conn.commit()

    # ---------- Resumen ----------
    print("\n=== RESUMEN ===")
    nprov = cursor.execute("SELECT COUNT(*) FROM proveedores").fetchone()[0]
    nins = cursor.execute("SELECT COUNT(*) FROM insumos").fetchone()[0]
    nrel = cursor.execute("SELECT COUNT(*) FROM proveedor_insumos").fetchone()[0]
    print(f"Proveedores en BD: {nprov}")
    print(f"Insumos en BD: {nins}")
    print(f"Relaciones proveedor-insumo en BD: {nrel}")
    if para_mensaje:
        print("\nSin proveedor resuelto:")
        print("\n".join(para_mensaje))
    print("\nImportación terminada.")


if __name__ == "__main__":
    main()

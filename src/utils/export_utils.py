import os
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import Qt
from PySide6.QtGui import QImage, QPageLayout, QPageSize, QPainter, QTextDocument
from PySide6.QtPrintSupport import QPrinter
from PySide6.QtWidgets import QFileDialog, QMessageBox, QTableWidget, QWidget


def export_table_to_excel(table: QTableWidget, titulo: str, parent: QWidget) -> Optional[str]:
    path, _ = QFileDialog.getSaveFileName(parent, f"Exportar {titulo}", f"{titulo}.xlsx",
                                           "Excel (*.xlsx)")
    if not path:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    visible_cols = [c for c in range(table.columnCount()) if not table.isColumnHidden(c)]
    headers = [table.horizontalHeaderItem(c).text() if table.horizontalHeaderItem(c) else ""
               for c in visible_cols]

    for ci, col_idx in enumerate(visible_cols):
        cell = ws.cell(row=1, column=ci + 1, value=headers[ci])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row in range(table.rowCount()):
        for ci, col_idx in enumerate(visible_cols):
            item = table.item(row, col_idx)
            val = item.text() if item else ""
            cell = ws.cell(row=row + 2, column=ci + 1, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if ci > 0 else "left")

    for ci in range(len(visible_cols)):
        ws.column_dimensions[get_column_letter(ci + 1)].width = max(12, len(headers[ci]) + 4)

    wb.save(path)
    return path


def print_table(table: QTableWidget, titulo: str, parent: QWidget) -> None:
    from src.components.preview_impresion import previsualizar_html
    html = _table_to_html(table, titulo)
    previsualizar_html(html, titulo=titulo, parent=parent)


def _table_to_html(table: QTableWidget, titulo: str) -> str:
    from src.utils.print_template import (
        _MENTA, _SALVIA, _SALVIA_OSCURA, _VERDE_OSCURO, _VERDE_MEDIO,
        logo_base64, nombre_empresa,
    )
    from datetime import datetime

    logo_b64 = logo_base64()
    empresa = nombre_empresa()
    ahora = datetime.now().strftime("%d/%m/%Y %H:%M")

    logo_html = ""
    if logo_b64:
        logo_html = (f'<img src="data:image/png;base64,{logo_b64}" '
                     'style="max-width:50px;max-height:50px;vertical-align:middle;margin-right:8px"/>')

    rows_html = ""
    visible_cols = [c for c in range(table.columnCount()) if not table.isColumnHidden(c)]
    headers = [table.horizontalHeaderItem(c).text() if table.horizontalHeaderItem(c) else ""
               for c in visible_cols]

    header_row = "".join(
        f"<th style='background:linear-gradient(135deg,{_SALVIA} 0%,{_SALVIA_OSCURA} 100%);"
        f"color:#fff;padding:7px 8px;text-align:center;font-size:10px;"
        f"border:1px solid {_SALVIA_OSCURA}'>{h}</th>"
        for h in headers)
    rows_html += f"<tr>{header_row}</tr>"

    for row_idx in range(table.rowCount()):
        cells = ""
        par = (row_idx % 2 == 1)
        bg = 'background:#f7faf6;' if par else ''
        for ci in visible_cols:
            item = table.item(row_idx, ci)
            val = item.text() if item else ""
            cells += (
                f"<td style='padding:6px;border:1px solid #e4e7e2;"
                f"font-size:10px;{bg}'>{val}</td>")
        rows_html += f"<tr>{cells}</tr>"

    return f"""<!DOCTYPE html>
<html><head><meta charset='utf-8'/><style>
@page {{ margin: 10mm; }}
body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; color: #374151; }}
.header {{
    background: linear-gradient(135deg, {_MENTA} 0%, {_SALVIA} 100%);
    padding: 12px 20px; border-bottom: 3px solid {_SALVIA_OSCURA};
}}
.header table {{ width: 100%; border-collapse: collapse; }}
.header td {{ vertical-align: middle; }}
.header .marca {{ font-size: 20px; font-weight: 800; color: {_VERDE_OSCURO}; letter-spacing: 2px; }}
.header .titulo {{ font-size: 16px; font-weight: 700; color: {_VERDE_OSCURO}; text-align: center; }}
.header .fecha {{ font-size: 11px; color: {_VERDE_MEDIO}; text-align: right; }}
.footer {{ margin-top: 14px; font-size: 10px; color: {_VERDE_MEDIO}; text-align: right; }}
</style></head><body>
<div class="header">
<table><tr>
<td style="width:35%">
  {logo_html}<span class="marca">{empresa.upper()}</span>
</td>
<td style="width:30%">
  <div class="titulo">{_esc(titulo)}</div>
</td>
<td style="width:35%">
  <div class="fecha">Generado el {ahora}</div>
</td>
</tr></table>
</div>
<table style='width:100%;border-collapse:collapse;margin-top:12px;font-family:Segoe UI,sans-serif'>
{rows_html}
</table>
<div class="footer">{empresa} - Generado el {ahora}</div>
</body></html>"""


def _logo_base64() -> str:
    """Logo del membrete en base64: prefiere `logonew.png` de la raíz del
    proyecto (membrete actual); si no existe, usa `views/assets/logo.png`."""
    raiz = Path(__file__).resolve().parent.parent.parent
    candidatos = [
        raiz / "logonew.png",
        Path(__file__).resolve().parent.parent / "views" / "assets" / "logo.png",
    ]
    import base64
    for ruta in candidatos:
        if ruta.exists():
            with open(str(ruta), "rb") as f:
                return base64.b64encode(f.read()).decode()
    return ""


def _nombre_empresa() -> str:
    try:
        from src.models.empresa_model import EmpresaModel
        return EmpresaModel().nombre_empresa()
    except Exception:
        return "GORETTI"


def _fmt_fecha(fecha: str) -> str:
    s = (fecha or "").strip()
    if not s:
        return ""
    fecha_solo = s.split(" ")[0]
    partes = fecha_solo.split("-")
    if len(partes) == 3 and len(partes[0]) == 4:
        anio, mes, dia = partes
        return f"{dia}/{mes}/{anio}"
    partes = fecha_solo.split("/")
    if len(partes) == 3:
        if len(partes[2]) == 4:
            dia, mes, anio = partes
        else:
            anio, mes, dia = partes[2], partes[1], partes[0]
        return f"{dia}/{mes}/{anio}"
    return fecha_solo


def _esc(texto: str) -> str:
    return (str(texto)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))


def _clave_numerica_talla(texto: str) -> float:
    try:
        return float(texto)
    except (TypeError, ValueError):
        return 0.0


def _oc_columnas_tallas(detalle: list[dict]) -> list[dict]:
    cols: list[dict] = []
    vistos: set[int] = set()
    for d in detalle:
        for t in d.get("tallas", []):
            tid = t["talla_id"]
            if tid not in vistos:
                vistos.add(tid)
                cols.append({"talla_id": tid, "talla": t.get("talla", "")})
    cols.sort(key=lambda t: _clave_numerica_talla(t.get("talla", "")))
    return cols


def _oc_subtotal_detalle(d: dict) -> float:
    """Subtotal del renglón: Σ(pares × precio) por talla si hay precios por talla;
    si no, cantidad × precio_unitario."""
    tallas = d.get("tallas", []) or []
    con_precio = [t for t in tallas if float(t.get("precio", 0) or 0) > 0]
    if con_precio:
        return sum(float(t.get("pares", 0) or 0) * float(t.get("precio", 0) or 0)
                   for t in con_precio)
    return float(d.get("cantidad", 0) or 0) * float(d.get("precio_unitario", 0) or 0)


def _oc_totales(detalle: list[dict], solo_remision: bool) -> tuple[float, float, float]:
    subtotal = sum(_oc_subtotal_detalle(d) for d in detalle)
    if solo_remision:
        iva = 0.0
    else:
        iva = round(subtotal * 0.16, 2)
    total = round(subtotal + iva, 2)
    return round(subtotal, 2), iva, total


def _oc_receipt_html(datos: dict, detalle: list[dict]) -> str:
    from datetime import datetime

    # Datos de empresa para el membrete
    try:
        from src.models.empresa_model import EmpresaModel
        em = EmpresaModel()
        emp_nombre = _esc(em.nombre_empresa().upper())
        emp_sub = _esc(em.razon_social())
    except Exception:
        emp_nombre = "GORETTI"
        emp_sub = ""

    solo_remision = bool(datos.get("solo_remision"))
    titulo = "REMI-SIÓN - ORDEN DE COMPRA" if solo_remision else \
        "RECIBO DE COMPRA - ORDEN DE COMPRA"
    subtotal, iva, total = _oc_totales(detalle, solo_remision)
    columnas = _oc_columnas_tallas(detalle)
    logo_b64 = _logo_base64()

    estatus = str(datos.get("estatus", "") or "").replace("_", " ").capitalize()
    proveedor = datos.get("proveedor_nombre") or "Compra a inventario"
    telefono = datos.get("proveedor_telefono") or ""
    email = datos.get("proveedor_email") or ""
    rfc = datos.get("proveedor_rfc") or ""
    direccion = datos.get("proveedor_direccion") or ""

    logo_html = ""
    if logo_b64:
        logo_html = (f'<img src="data:image/jpeg;base64,{logo_b64}" '
                     f'style="max-width:90px;max-height:90px;vertical-align:middle;margin-right:8px"/>')

    th_tallas = "".join(
        f"<th>{_esc('#')}{_esc(c['talla'])}</th>" for c in columnas)

    rows = ""
    for d in detalle:
        por_talla = {int(t["talla_id"]): int(t.get("pares", 0) or 0)
                     for t in d.get("tallas", [])}
        celdas = "".join(
            f"<td class='td-num'>{por_talla.get(int(c['talla_id']), 0) or ''}</td>"
            for c in columnas)
        cant = int(d.get("cantidad", 0) or 0)
        precio = float(d.get("precio_unitario", 0) or 0)
        sub = _oc_subtotal_detalle(d)
        rows += f"""<tr>
            <td style='text-align:left'>{_esc(d.get('insumo_nombre', ''))}</td>
            {celdas}
            <td class='td-num'>{cant}</td>
            <td class='td-der'>${precio:,.2f}</td>
            <td class='td-der'>${sub:,.2f}</td>
        </tr>"""

    filas_iva = ""
    if not solo_remision:
        filas_iva = f"""<tr>
            <td class='lbl'>IVA (16%)</td>
            <td class='val'>${iva:,.2f}</td>
        </tr>"""

    datos_proveedor = [l for l in [
        _esc(proveedor),
        (f"Tel: {_esc(telefono)}" if telefono else ""),
        (f"Email: {_esc(email)}" if email else ""),
        (f"RFC: {_esc(rfc)}" if rfc else ""),
        (f"Dirección: {_esc(direccion)}" if direccion else ""),
    ] if l]

    vendido_html = "".join(
        f"<div class='{'empresa' if i == 0 else 'sub'}'>"
        f"{l}</div>" for i, l in enumerate(datos_proveedor))

    info_pago = [l for l in [
        _esc(proveedor),
        (f"Tel: {_esc(telefono)}" if telefono else ""),
        (f"Email: {_esc(email)}" if email else ""),
    ] if l]
    info_pago_html = "".join(f"<div>{l}</div>" for l in info_pago)

    # Observaciones
    obs_texto = str(datos.get("observaciones") or "").strip()
    if obs_texto:
        obs_html = (
            f"<div class='obs'>"
            f"<div class='lbl'>Observaciones</div>"
            f"<div>{_esc(obs_texto).replace(chr(10), '<br/>')}</div>"
            f"</div>")
    else:
        obs_html = ""

    ahora = datetime.now().strftime("%d/%m/%Y %H:%M")
    sub_label = emp_sub if emp_sub and emp_sub != emp_nombre else ""

    from src.utils.print_template import (
        _MENTA, _SALVIA, _SALVIA_OSCURA, _VERDE_OSCURO, _VERDE_MEDIO, _BLANCO,
    )

    return f"""<!DOCTYPE html>
<html><head><meta charset='utf-8'/><style>
@page {{ margin: 14mm; }}
body {{ font-family: Segoe UI, Arial, sans-serif; font-size: 12px; color: #374151; margin: 0; }}
.encabezado {{ background: linear-gradient(135deg, {_MENTA} 0%, {_SALVIA} 100%);
    border-bottom: 3px solid {_SALVIA_OSCURA}; padding-bottom: 10px; }}
.encabezado table {{ width: 100%; border-collapse: collapse; }}
.encabezado td {{ vertical-align: middle; }}
.marca {{ font-size: 22px; font-weight: 800; color: {_VERDE_OSCURO}; letter-spacing: 2px; }}
.razon {{ font-size: 11px; color: {_VERDE_MEDIO}; margin-top: 2px; }}
.datos {{ font-size: 10px; color: {_VERDE_MEDIO}; margin-top: 2px; }}
.titulo {{ font-size: 17px; font-weight: 700; color: {_VERDE_OSCURO}; text-align: center; }}
.no-fecha {{ text-align: right; font-size: 12px; color: #374151; }}
.no-fecha .num {{ font-size: 15px; font-weight: bold; color: {_VERDE_OSCURO}; }}
.vendido {{ border: 1px solid {_SALVIA_OSCURA}; border-left: 4px solid {_SALVIA}; padding: 10px 14px; margin-top: 14px; background: {_BLANCO}; }}
.vendido .lbl {{ font-weight: bold; color: {_VERDE_MEDIO}; font-size: 11px; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 4px; }}
.vendido .empresa {{ font-size: 15px; font-weight: bold; color: {_VERDE_OSCURO}; }}
.vendido .sub {{ font-size: 12px; color: {_VERDE_MEDIO}; }}
table.items {{ width: 100%; border-collapse: collapse; margin-top: 14px; }}
table.items th {{ background: linear-gradient(135deg, {_SALVIA} 0%, {_SALVIA_OSCURA} 100%);
    color: #ffffff; padding: 7px 6px; font-size: 10px; text-align: center;
    border: 1px solid {_SALVIA_OSCURA}; }}
table.items td {{ border: 1px solid #e4e7e2; padding: 6px; font-size: 11px; background: {_BLANCO}; }}
table.items tr:nth-child(even) td {{ background: #f7faf6; }}
.td-num {{ text-align: center; }}
.td-der {{ text-align: right; }}
.resumen {{ margin-top: 14px; margin-left: auto; width: 250px; border-collapse: collapse; }}
.resumen td {{ padding: 5px 10px; font-size: 12px; border: 1px solid #e4e7e2; }}
.resumen .lbl {{ color: {_VERDE_MEDIO}; }}
.resumen .val {{ text-align: right; font-weight: bold; }}
.resumen .fila-total td {{ background: linear-gradient(135deg, {_SALVIA} 0%, {_SALVIA_OSCURA} 100%);
    color: #ffffff; font-size: 15px; font-weight: bold; border-color: {_SALVIA_OSCURA}; }}
.metodo {{ margin-top: 16px; font-size: 13px; }}
.metodo b {{ color: {_VERDE_OSCURO}; }}
.pago {{ border: 1px solid {_SALVIA_OSCURA}; border-left: 4px solid {_SALVIA}; padding: 10px 14px; margin-top: 10px; font-size: 12px; background: {_BLANCO}; }}
.pago .lbl, .obs .lbl {{ font-weight: bold; color: {_VERDE_MEDIO}; font-size: 11px; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 4px; }}
.obs {{ border: 1px solid {_SALVIA_OSCURA}; border-left: 4px solid {_SALVIA}; padding: 10px 14px; margin-top: 14px; font-size: 11px; background: {_BLANCO}; }}
.pie {{ margin-top: 26px; padding-top: 10px; border-top: 2px solid {_SALVIA_OSCURA}; text-align: center; }}
.pie .gracias {{ font-size: 14px; font-weight: bold; color: {_VERDE_OSCURO}; }}
.pie .marca {{ font-size: 16px; font-weight: bold; color: {_VERDE_OSCURO}; letter-spacing: 2px; margin-top: 4px; }}
.pie .leyenda {{ font-size: 9px; color: {_VERDE_MEDIO}; margin-top: 4px; }}
</style></head><body>
<div class='encabezado'>
<table><tr>
<td style='width:35%'>
  {logo_html}<span class='marca'>{emp_nombre}</span>
  {'<div class="razon">' + sub_label + '</div>' if sub_label else ''}
</td>
<td style='width:30%;text-align:center'>
  <div class='titulo'>{titulo}</div>
</td>
<td class='no-fecha' style='width:35%'>
  <div>NO. <span class='num'>{_esc(datos.get('folio', ''))}</span></div>
  <div>FECHA: <b>{_fmt_fecha(datos.get('fecha_emision', ''))}</b></div>
  <div class='datos'>Estatus: {_esc(estatus)}</div>
</td>
</tr></table>
</div>

<div class='vendido'>
  <div class='lbl'>Vendido a:</div>
  {vendido_html}
</div>

<table class='items'>
<tr>
  <th style='text-align:left;min-width:170px'>NOMBRE</th>
  {th_tallas}
  <th>TOTAL PARES</th>
  <th>VALOR UNITARIO</th>
  <th>TOTAL</th>
</tr>
{rows}
</table>

<table class='resumen'>
<tr><td class='lbl'>Subtotal</td><td class='val'>${subtotal:,.2f}</td></tr>
{filas_iva}
<tr class='fila-total'><td>TOTAL</td><td>${total:,.2f}</td></tr>
</table>

{obs_html}
<div class='metodo'>Método de pago: <b>{_esc(datos.get('metodo_pago') or 'Transferencia bancaria')}</b></div>

<div class='pago'>
  <div class='lbl'>Información de Pago</div>
  {info_pago_html}
</div>

<div class='pie'>
  <div class='gracias'>Gracias por su compra.</div>
  <div class='marca'>{emp_nombre}</div>
  <div class='leyenda'>Generado por {emp_nombre} el {ahora}</div>
</div>

</body></html>"""


def print_orden_compra(datos: dict, detalle: list[dict], parent: QWidget) -> None:
    path, _ = QFileDialog.getSaveFileName(
        parent, "Guardar PDF - Orden de Compra",
        f"OC_{datos.get('folio', '')}.pdf", "PDF (*.pdf)")
    if not path:
        return

    doc = QTextDocument()
    doc.setHtml(_oc_receipt_html(datos, detalle))
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setPageSize(QPageSize(QPageSize.Letter))
    printer.setPageOrientation(
        QPageLayout.Orientation.Landscape
        if len(_oc_columnas_tallas(detalle)) > 6
        else QPageLayout.Orientation.Portrait)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(path)
    doc.print_(printer)


def export_orden_compra_excel(datos: dict, detalle: list[dict], parent: QWidget) -> Optional[str]:
    path, _ = QFileDialog.getSaveFileName(
        parent, "Exportar Excel - Orden de Compra",
        f"OC_{datos.get('folio', '')}.xlsx", "Excel (*.xlsx)")
    if not path:
        return None
    _write_oc_excel(path, datos, detalle)
    return path


def _write_oc_excel(path: str, datos: dict, detalle: list[dict]) -> None:
    from datetime import datetime

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    solo_remision = bool(datos.get("solo_remision"))
    titulo = "REMI-SIÃ“N - ORDEN DE COMPRA" if solo_remision else \
        "RECIBO DE COMPRA - ORDEN DE COMPRA"
    columnas = _oc_columnas_tallas(detalle)
    subtotal, iva, total = _oc_totales(detalle, solo_remision)

    n_tallas = len(columnas)
    n_cols = 1 + n_tallas + 3

    navy = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    light = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    blanco = Font(color="FFFFFF")
    bold = Font(bold=True)
    thin = Side(style="thin", color="E2E8F0")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    centro = Alignment(horizontal="center", vertical="center")
    der = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "OC"
    last = n_cols

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = Font(bold=True, size=16, color="1D4ED8")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last)
    c = ws.cell(row=2, column=1,
                value=f"NO. {datos.get('folio', '')}    FECHA: {_fmt_fecha(datos.get('fecha_emision', ''))}")
    c.font = Font(size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")

    fila = 4
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
    c = ws.cell(row=fila, column=1, value="Vendido a:")
    c.font = Font(bold=True, size=10, color="475569")
    fila += 1

    proveedor = datos.get("proveedor_nombre") or "Compra a inventario"
    lineas = [
        ("empresa", proveedor),
        ("sub", f"Tel: {datos.get('proveedor_telefono') or 'â€”'}"),
        ("sub", f"Email: {datos.get('proveedor_email') or 'â€”'}"),
        ("sub", f"RFC: {datos.get('proveedor_rfc') or 'â€”'}"),
        ("sub", f"DirecciÃ³n: {datos.get('proveedor_direccion') or 'â€”'}"),
    ]
    for tipo, texto in lineas:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
        c = ws.cell(row=fila, column=1, value=texto)
        c.font = Font(bold=(tipo == "empresa"), size=13 if tipo == "empresa" else 11,
                      color="1F2937" if tipo == "empresa" else "64748B")
        fila += 1

    fila += 1
    header_fila = fila
    ws.cell(row=fila, column=1, value="NOMBRE")
    for i, col in enumerate(columnas):
        ws.cell(row=fila, column=2 + i, value=f"#{col['talla']}")
    ws.cell(row=fila, column=1 + n_tallas + 1, value="TOTAL PARES")
    ws.cell(row=fila, column=1 + n_tallas + 2, value="VALOR UNITARIO")
    ws.cell(row=fila, column=1 + n_tallas + 3, value="TOTAL")
    for col in range(1, last + 1):
        cc = ws.cell(row=fila, column=col)
        cc.font = Font(bold=True, color="FFFFFF", size=10)
        cc.fill = navy
        cc.alignment = centro
        cc.border = borde
    ws.row_dimensions[fila].height = 22
    fila += 1

    for d in detalle:
        por_talla = {int(t["talla_id"]): int(t.get("pares", 0) or 0)
                     for t in d.get("tallas", [])}
        ws.cell(row=fila, column=1, value=d.get("insumo_nombre", ""))
        for i, col in enumerate(columnas):
            ws.cell(row=fila, column=2 + i,
                    value=por_talla.get(int(col["talla_id"]), 0) or "")
        cant = int(d.get("cantidad", 0) or 0)
        precio = float(d.get("precio_unitario", 0) or 0)
        ws.cell(row=fila, column=1 + n_tallas + 1, value=cant)
        ws.cell(row=fila, column=1 + n_tallas + 2, value=precio)
        ws.cell(row=fila, column=1 + n_tallas + 3, value=round(_oc_subtotal_detalle(d), 2))
        for col in range(1, last + 1):
            cc = ws.cell(row=fila, column=col)
            cc.border = borde
            if col > 1:
                cc.alignment = centro if col <= 1 + n_tallas else der
            cc.font = Font(size=10)
        if fila % 2 == 0:
            for col in range(1, last + 1):
                ws.cell(row=fila, column=col).fill = light
        fila += 1

    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=max(1, last - 3))
    c = ws.cell(row=fila, column=1, value="Subtotal")
    c.font = Font(size=11)
    c.alignment = der
    c2 = ws.cell(row=fila, column=last - 2 if last - 2 > 1 else last, value=round(subtotal, 2))
    c2.font = Font(bold=True, size=11)
    c2.alignment = der
    fila += 1

    if not solo_remision:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=max(1, last - 3))
        c = ws.cell(row=fila, column=1, value="IVA (16%)")
        c.font = Font(size=11)
        c.alignment = der
        c2 = ws.cell(row=fila, column=last - 2 if last - 2 > 1 else last, value=iva)
        c2.font = Font(bold=True, size=11)
        c2.alignment = der
        fila += 1

    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=max(1, last - 3))
    c = ws.cell(row=fila, column=1, value="TOTAL")
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = navy
    c.alignment = der
    c2 = ws.cell(row=fila, column=last - 2 if last - 2 > 1 else last, value=total)
    c2.font = Font(bold=True, size=14, color="FFFFFF")
    c2.fill = navy
    c2.alignment = der
    fila += 2

    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
    c = ws.cell(row=fila, column=1,
                value=f"MÃ©todo de pago: {datos.get('metodo_pago') or 'Transferencia bancaria'}")
    c.font = Font(bold=True, size=12, color="1D4ED8")
    fila += 2

    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
    c = ws.cell(row=fila, column=1, value="InformaciÃ³n de Pago")
    c.font = Font(bold=True, size=10, color="475569")
    fila += 1
    for texto in [
        proveedor,
        f"Tel: {datos.get('proveedor_telefono') or 'â€”'}",
        f"Email: {datos.get('proveedor_email') or 'â€”'}",
    ]:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
        c = ws.cell(row=fila, column=1, value=texto)
        c.font = Font(size=11)
        fila += 1

    fila += 2
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
    c = ws.cell(row=fila, column=1, value="Gracias por su compra.")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
    c = ws.cell(row=fila, column=1, value=_nombre_empresa().upper())
    c.font = Font(bold=True, size=16, color="1D4ED8")
    c.alignment = Alignment(horizontal="center")
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
    c = ws.cell(row=fila, column=1,
                value=f"Generado por {_nombre_empresa()} el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = Font(size=9, color="94A3B8")
    c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 42
    for i in range(n_tallas):
        ws.column_dimensions[get_column_letter(2 + i)].width = 9
    ws.column_dimensions[get_column_letter(1 + n_tallas + 1)].width = 12
    ws.column_dimensions[get_column_letter(1 + n_tallas + 2)].width = 15
    ws.column_dimensions[get_column_letter(1 + n_tallas + 3)].width = 13

    wb.save(path)


def _pedido_columnas_puntos(detalle: list[dict]) -> list[dict]:
    return _oc_columnas_tallas(detalle)


def _pedido_totales(detalle: list[dict]) -> int:
    return sum(int(p.get("pares", 0) or 0)
               for d in detalle for p in d.get("puntos", []))


def print_pedido_cliente(datos: dict, detalle: list[dict], parent: QWidget) -> None:
    from src.components.preview_impresion import previsualizar_html
    folio = datos.get('folio', '')
    titulo = f"Pedido de Cliente {folio}" if folio else "Pedido de Cliente"
    previsualizar_html(_pedido_html(datos, detalle), titulo=titulo, parent=parent)


def _pedido_html(datos: dict, detalle: list[dict]) -> str:
    """Pedido de cliente con plantilla mint/salvia."""
    from src.utils.print_template import (
        esc as t_esc, fmt_fecha as t_fmt_fecha,
        html_cabecera, html_ondas_superiores, html_obs_bloque,
        html_pie, wrap_hoja,
    )

    columnas = _pedido_columnas_puntos(detalle)
    total = _pedido_totales(detalle)

    estatus = str(datos.get("estatus", "") or "").replace("_", " ").capitalize()
    cliente = datos.get("cliente_nombre", "")
    telefono = datos.get("cliente_telefono", "")
    email = datos.get("cliente_email", "")
    dir_envio = datos.get("direccion_envio", "")
    folio_pedido = datos.get("folio_pedido", "")
    suela = datos.get("suela", "")
    horma = datos.get("horma", "")

    th_tallas = "".join(
        f"<th>{t_esc(c['talla'])}</th>" for c in columnas)

    rows = ""
    for d in detalle:
        por_talla = {}
        for t in d.get("tallas", []):
            por_talla[str(t.get("talla", ""))] = int(t.get("pares", 0) or 0)
        celdas = "".join(
            f"<td>{por_talla.get(c['talla'], 0) or ''}</td>"
            for c in columnas)
        cant = int(d.get("total_pares", 0) or 0)
        rows += f"""<tr>
            <td style='text-align:left'>{t_esc(d.get('modelo_nombre', ''))}</td>
            {celdas}
            <td style='font-weight:700'>{cant}</td>
        </tr>"""

    lineas_cliente = [l for l in [
        t_esc(cliente),
        (f"Tel: {t_esc(telefono)}" if telefono else ""),
        (f"Email: {t_esc(email)}" if email else ""),
        (f"Envio: {t_esc(dir_envio)}" if dir_envio else ""),
    ] if l]
    cliente_html = "".join(
        f"<div style='{'font-size:15px;font-weight:700;color:#2f4f3a' if i == 0 else 'color:#5b6b60;font-size:12px'}'>"
        f"{l}</div>" for i, l in enumerate(lineas_cliente))

    cabecera = html_cabecera(
        "PEDIDO DE CLIENTE", folio=t_esc(datos.get("folio", "")),
        fecha=datos.get("fecha_pedido", ""),
        extra_derecha=f"Estatus: {t_esc(estatus)}")

    contenido = f"""
{cabecera}
{html_ondas_superiores()}

<table width="100%" cellpadding="0" cellspacing="0" style="margin:0;">
<tr><td style="padding:0 14mm;">

<div class='bloque'>
  <div class='lbl'>Cliente:</div>
  {cliente_html}
</div>

<table class='items' width="100%" cellpadding="0" cellspacing="0">
<tr>
  <th style='min-width:170px'>MODELO</th>
  {th_tallas}
  <th>TOTAL PARES</th>
</tr>
{rows}
</table>

<table class='resumen'>
<tr class='fila-total'><td>TOTAL PARES</td><td>{total}</td></tr>
</table>

</td></tr></table>

{html_pie("Pedido de Cliente")}
"""
    return wrap_hoja(contenido)


def export_pedido_cliente_excel(datos: dict, detalle: list[dict], parent: QWidget) -> Optional[str]:
    path, _ = QFileDialog.getSaveFileName(
        parent, "Exportar Excel - Pedido de Cliente",
        f"PED_{datos.get('folio', '')}.xlsx", "Excel (*.xlsx)")
    if not path:
        return None
    _write_pedido_excel(path, datos, detalle)
    return path


def exportar_programacion_excel(lineas: list[dict], titulo: str,
                                incluir_semana: bool = False,
                                parent: QWidget | None = None,
                                grupos: list | None = None) -> Optional[str]:
    """Exporta la programación a Excel replicando el formato del reporte que
    imprime (generar_html_programacion): título morado sobre rosa, encabezado
    morado, una columna por talla y fila final de totales en negrita.

    `grupos`: lista opcional de (etiqueta, [lineas]) para insertar filas de
    grupo moradas (segundo encabezado) cuando la tabla está agrupada."""
    nombre = "".join(c for c in titulo if c.isalnum() or c in " _-") or "Programacion"
    path, _ = QFileDialog.getSaveFileName(
        parent, "Exportar Excel - Programación", f"{nombre}.xlsx", "Excel (*.xlsx)")
    if not path:
        return None
    _write_programacion_excel(path, lineas, titulo, incluir_semana, grupos)
    return path


def _write_programacion_excel(path: str, lineas: list[dict], titulo: str,
                              incluir_semana: bool, grupos: list | None) -> None:
    from datetime import datetime

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    def tallas_ordenadas(ls: list[dict]) -> list[str]:
        tallas: list[str] = []
        vistos: set[str] = set()
        for linea in ls:
            for t in linea.get("tallas") or []:
                talla = str(t.get("talla", "") or "").strip()
                if talla and talla not in vistos:
                    vistos.add(talla)
                    tallas.append(talla)
        tallas.sort(key=lambda x: (float(x), x))
        return tallas

    tallas = tallas_ordenadas(lineas)
    total_por_talla: dict[str, int] = {talla: 0 for talla in tallas}
    gran_total = 0

    fijas = [
        ("cliente", "CLIENTE"),
        ("folio_prog", "FOLIO PROG."),
        ("folio_pedido", "FOLIO PEDIDO"),
        ("modelo", "MODELO"),
        ("piel", "PIEL"),
        ("color", "COLOR"),
        ("fecha_prog", "FECHA PROG."),
    ]

    n_texto = len(fijas) + (1 if incluir_semana else 0)
    n_cols = n_texto + len(tallas) + 1

    morado = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    rosa_total = PatternFill(start_color="FFE6FF", end_color="FFE6FF", fill_type="solid")
    grupo_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    par_fill = PatternFill(start_color="FDF2FF", end_color="FDF2FF", fill_type="solid")
    blanco = Font(color="FFFFFF", bold=True)
    negrita = Font(bold=True)
    thin = Side(style="thin", color="D946EF")
    borde_total = Border(left=thin, right=thin, top=thin, bottom=thin)
    fina = Side(style="thin", color="E5E7EB")
    borde = Border(left=fina, right=fina, top=fina, bottom=fina)
    centro = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "PROGRAMACIÓN"
    last = n_cols

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = Font(bold=True, size=15, color="7C3AED")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last)
    c = ws.cell(row=2, column=1,
                value=f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = Font(size=10, color="6B7280")
    c.alignment = Alignment(horizontal="center", vertical="center")

    fila = 4
    encabezados = []
    if incluir_semana:
        encabezados.append("SEMANA")
    encabezados += [t for _, t in fijas]
    encabezados += tallas
    encabezados.append("TOTAL PARES")
    for j, h in enumerate(encabezados):
        cc = ws.cell(row=fila, column=j + 1, value=h)
        cc.font = blanco
        cc.fill = morado
        cc.alignment = centro
        cc.border = borde
    ws.row_dimensions[fila].height = 20
    fila += 1

    def escribir_linea(linea: dict) -> None:
        nonlocal fila, gran_total
        texto = []
        if incluir_semana:
            texto.append(linea.get("semana", ""))
        base = {
            "cliente": linea.get("cliente", ""),
            "folio_prog": linea.get("folio_prog", ""),
            "folio_pedido": linea.get("folio_pedido", ""),
            "modelo": linea.get("modelo", ""),
            "piel": linea.get("piel", ""),
            "color": linea.get("color", ""),
            "fecha_prog": linea.get("fecha_prog", "") or "",
        }
        texto += [base.get(key, "") for key, _ in fijas]
        numeros = []
        por_talla = {str(t.get("talla", "")): int(t.get("pares", 0) or 0)
                     for t in linea.get("tallas") or []}
        for talla in tallas:
            pares = por_talla.get(talla, 0)
            total_por_talla[talla] += pares
            numeros.append(str(pares or ""))
        total = int(linea.get("total_pares", 0) or 0)
        gran_total += total
        numeros.append(str(total))
        valores = texto + numeros
        for j, v in enumerate(valores):
            cc = ws.cell(row=fila, column=j + 1, value=v)
            cc.border = borde
            cc.alignment = centro if j >= n_texto else Alignment(horizontal="left",
                                                                 vertical="center")
            cc.font = Font(size=10)
        if fila % 2 == 0:
            for j in range(1, last + 1):
                ws.cell(row=fila, column=j).fill = par_fill
        fila += 1

    def escribir_grupo(etiqueta: str) -> None:
        nonlocal fila
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
        cc = ws.cell(row=fila, column=1, value=etiqueta)
        cc.font = blanco
        cc.fill = grupo_fill
        cc.alignment = Alignment(horizontal="left", vertical="center")
        fila += 1

    if grupos:
        for etiqueta, ls in grupos:
            escribir_grupo(etiqueta)
            for linea in ls:
                escribir_linea(linea)
    else:
        for linea in lineas:
            escribir_linea(linea)

    fila_total = []
    if incluir_semana:
        fila_total.append("")
    fila_total.append("TOTAL")
    fila_total += [""] * (len(fijas) - 1)
    fila_total += [str(total_por_talla.get(talla, 0)) for talla in tallas]
    fila_total.append(str(gran_total))
    for j, v in enumerate(fila_total):
        cc = ws.cell(row=fila, column=j + 1, value=v)
        cc.font = negrita
        cc.fill = rosa_total
        cc.border = borde_total
        cc.alignment = centro if j >= n_texto else Alignment(horizontal="left",
                                                             vertical="center")
    fila += 1

    ws.column_dimensions["A"].width = 34
    for i in range(1, n_texto):
        ws.column_dimensions[get_column_letter(i + 1)].width = 16
    for i in range(len(tallas)):
        ws.column_dimensions[get_column_letter(n_texto + 1 + i)].width = 9
    ws.column_dimensions[get_column_letter(last)].width = 12

    wb.save(path)


def _write_pedido_excel(path: str, datos: dict, detalle: list[dict]) -> None:
    from datetime import datetime

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    columnas = _oc_columnas_tallas(detalle)
    total = _pedido_totales(detalle)

    n_tallas = len(columnas)
    n_cols = 1 + 3 + n_tallas + 1

    navy = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    light = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    blanco = Font(color="FFFFFF")
    bold = Font(bold=True)
    thin = Side(style="thin", color="E2E8F0")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)
    centro = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "PEDIDO"
    last = n_cols

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last)
    c = ws.cell(row=1, column=1, value="PEDIDO DE CLIENTE")
    c.font = Font(bold=True, size=16, color="1D4ED8")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    fecha_prog = f"   FECHA PROGRAMADO: {_fmt_fecha(datos.get('fecha_programado', ''))}"
    folio_pedido = datos.get("folio_pedido") or ""
    suela = datos.get("suela") or ""
    horma = datos.get("horma") or ""
    extras = " ".join(x for x in [
        f"FOLIO PEDIDO: {folio_pedido}" if folio_pedido else "",
        f"SUELA: {suela}" if suela else "",
        f"HORMA: {horma}" if horma else "",
    ] if x)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last)
    c = ws.cell(row=2, column=1,
                value=f"NO. {datos.get('folio', '')}    FECHA PEDIDO: {_fmt_fecha(datos.get('fecha_pedido', ''))}"
                      f"{fecha_prog}    {extras}    ESTATUS: {str(datos.get('estatus', '') or '').replace('_', ' ').capitalize()}".replace("    ", " ").strip())
    c.font = Font(size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")

    fila = 4
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
    c = ws.cell(row=fila, column=1, value="Cliente:")
    c.font = Font(bold=True, size=10, color="475569")
    fila += 1

    cliente = datos.get("cliente_nombre") or ""
    lineas = [
        ("empresa", cliente),
        ("sub", f"Tel: {datos.get('cliente_telefono') or 'â€”'}"),
        ("sub", f"Email: {datos.get('cliente_email') or 'â€”'}"),
        ("sub", f"RFC: {datos.get('cliente_rfc') or 'â€”'}"),
        ("sub", f"DirecciÃ³n: {datos.get('cliente_direccion') or 'â€”'}"),
    ]
    for tipo, texto in lineas:
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
        c = ws.cell(row=fila, column=1, value=texto)
        c.font = Font(bold=(tipo == "empresa"), size=13 if tipo == "empresa" else 11,
                      color="1F2937" if tipo == "empresa" else "64748B")
        fila += 1

    fila += 1
    header_fila = fila
    ws.cell(row=fila, column=1, value="MODELO")
    ws.cell(row=fila, column=2, value="PIEL")
    ws.cell(row=fila, column=3, value="COLOR")
    for i, col in enumerate(columnas):
        ws.cell(row=fila, column=4 + i, value=f"#{col['punto']}")
    ws.cell(row=fila, column=4 + n_tallas, value="TOTAL PARES")
    for col in range(1, last + 1):
        cc = ws.cell(row=fila, column=col)
        cc.font = Font(bold=True, color="FFFFFF", size=10)
        cc.fill = navy
        cc.alignment = centro
        cc.border = borde
    ws.row_dimensions[fila].height = 22
    fila += 1

    for d in detalle:
        por_talla = {int(t["punto_id"]): int(t.get("pares", 0) or 0)
                     for t in d.get("puntos", [])}
        ws.cell(row=fila, column=1, value=d.get("modelo", ""))
        ws.cell(row=fila, column=2, value=d.get("piel", ""))
        ws.cell(row=fila, column=3, value=d.get("color", ""))
        for i, col in enumerate(columnas):
            ws.cell(row=fila, column=4 + i,
                    value=por_talla.get(int(col["punto_id"]), 0) or "")
        ws.cell(row=fila, column=4 + n_tallas, value=sum(por_talla.values()))
        for col in range(1, last + 1):
            cc = ws.cell(row=fila, column=col)
            cc.border = borde
            if col > 1:
                cc.alignment = centro
            cc.font = Font(size=10)
        if fila % 2 == 0:
            for col in range(1, last + 1):
                ws.cell(row=fila, column=col).fill = light
        fila += 1

    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last - 1)
    c = ws.cell(row=fila, column=1, value="TOTAL PARES")
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = navy
    c.alignment = Alignment(horizontal="right", vertical="center")
    c2 = ws.cell(row=fila, column=last, value=total)
    c2.font = Font(bold=True, size=14, color="FFFFFF")
    c2.fill = navy
    c2.alignment = centro
    fila += 2

    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
    c = ws.cell(row=fila, column=1, value="Gracias por su compra.")
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
    c = ws.cell(row=fila, column=1, value=_nombre_empresa().upper())
    c.font = Font(bold=True, size=16, color="1D4ED8")
    c.alignment = Alignment(horizontal="center")
    fila += 1
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=last)
    c = ws.cell(row=fila, column=1,
                value=f"Generado por {_nombre_empresa()} el {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.font = Font(size=9, color="94A3B8")
    c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    for i in range(n_tallas):
        ws.column_dimensions[get_column_letter(4 + i)].width = 9
    ws.column_dimensions[get_column_letter(4 + n_tallas)].width = 13

    wb.save(path)

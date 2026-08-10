"""Componente propio aprobado: ComplexGrid.

Tabla de datos con búsqueda, filtros, agrupación, vistas lista/iconos/tabla,
acciones por registro y exportación/impresión. Aprobado desde el Sandbox.

Configuración (desde código):
    set_columnas([{"key", "titulo", "ancho", "tipo"}])
    set_datos(registros)
    set_renderers(fila, claves, tarjeta, lista)
    set_acciones([{"texto", "icono", "color", "callback"}])
    set_filtros([fn(rec) -> bool])
    set_agrupacion(clave_columna | None)
    set_plantilla_excel(ruta, inicio="A3")
    set_reporte_config({...})
"""

from functools import partial
from pathlib import Path

from openpyxl import load_workbook
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QFont, QTextDocument
from PySide6.QtPrintSupport import QPrinter
from PySide6.QtWidgets import (
    QButtonGroup, QComboBox, QFileDialog, QGridLayout, QHBoxLayout, QHeaderView,
    QLabel, QLineEdit, QListWidget, QListWidgetItem, QPushButton, QScrollArea,
    QStackedWidget, QTableWidget, QTableWidgetItem, QToolButton, QVBoxLayout,
    QWidget,
)

from src.utils.export_utils import export_table_to_excel, print_table
from src.utils.icons import mono_icon
from src.utils.odoo_list import _ItemOrdenable, _Tarjeta
from src.utils.table_utils import configurar_tabla_excel


def _esc(texto: str) -> str:
    return (texto.replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;"))


def _celda_inicio(texto: str) -> tuple[int, int]:
    texto = texto.strip().upper()
    col = 0
    i = 0
    while i < len(texto) and texto[i].isalpha():
        col = col * 26 + (ord(texto[i]) - 64)
        i += 1
    fila = int(texto[i:]) if texto[i:] else 1
    return fila, col


class ComplexGrid(QWidget):
    """Tabla de datos con búsqueda, filtros, agrupación, vistas y acciones.

    Configuración principal (desde código):
        set_columnas([{"key", "titulo", "ancho", "tipo"}])
        set_datos(registros)
        set_renderers(fila, claves, tarjeta, lista)
        set_acciones([{"texto", "icono", "color", "callback"}])
        set_filtros([fn(rec) -> bool])
        set_agrupacion(clave_columna | None)
        set_plantilla_excel(ruta, inicio="A3")
        set_reporte_config({...})
    """

    doubleClicked = Signal()
    selectionChanged = Signal()

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._col_config: list[dict] = []
        self._registros: list = []
        self._visibles: list = []
        self._fila_fn = None
        self._claves_fn = None
        self._tarjeta_fn = None
        self._lista_fn = None
        self._acciones: list[dict] = []
        self._filtros: list = []
        self._buscar = ""
        self._agrupar_por: str | None = None
        self._plantilla_excel: str | None = None
        self._plantilla_inicio = "A3"
        self._reporte_cfg: dict = {}
        self._sort_col = -1
        self._sort_asc = True
        self._mapa_rec: list = []
        self._idx_lista = 0
        self._idx_tabla = 1
        self._idx_iconos = 2
        self._setup_ui()

    # ------------------------------------------------------------------ UI
    def _setup_ui(self) -> None:
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(8)

        bar = QHBoxLayout()
        bar.setSpacing(8)

        self._txt_buscar = QLineEdit()
        self._txt_buscar.setPlaceholderText("Buscar...")
        self._txt_buscar.setClearButtonEnabled(True)
        self._txt_buscar.setMinimumWidth(220)
        self._txt_buscar.textChanged.connect(self._on_buscar)
        bar.addWidget(self._txt_buscar)

        self._lbl_estado = QLabel("")
        self._lbl_estado.setObjectName("sectionSubtitle")
        bar.addWidget(self._lbl_estado)

        self._cmb_agrupar = QComboBox()
        self._cmb_agrupar.addItem("Sin agrupar", None)
        self._cmb_agrupar.currentIndexChanged.connect(self._on_cmb_agrupar)
        bar.addWidget(self._cmb_agrupar)

        bar.addStretch()

        self._btn_excel = self._crear_boton("exportar", "Excel", "#16a34a",
                                            self.exportar_excel)
        self._btn_pdf = self._crear_boton("pdf", "PDF", "#dc2626",
                                          self.exportar_pdf)
        self._btn_imprimir = self._crear_boton("imprimir", "Imprimir", "#4f46e5",
                                               self.imprimir)
        for b in (self._btn_excel, self._btn_pdf, self._btn_imprimir):
            bar.addWidget(b)

        self._btn_lista = self._crear_boton_vista("lista", "Lista")
        self._btn_tabla = self._crear_boton_vista("tabla", "Tabla")
        self._btn_iconos = self._crear_boton_vista("iconos", "Iconos")
        grupo = QButtonGroup(self)
        grupo.setExclusive(True)
        for btn in (self._btn_lista, self._btn_tabla, self._btn_iconos):
            grupo.addButton(btn)
            bar.addWidget(btn)
        lay.addLayout(bar)

        self._stack = QStackedWidget()
        self._pag_lista = QListWidget()
        self._pag_lista.itemSelectionChanged.connect(self._sincronizar_seleccion)
        self._pag_lista.itemDoubleClicked.connect(lambda _it: self._emit_doble())
        self._stack.addWidget(self._pag_lista)

        self._pag_tabla = QTableWidget()
        self._pag_tabla.setSelectionBehavior(QTableWidget.SelectRows)
        self._pag_tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        self._pag_tabla.setAlternatingRowColors(True)
        configurar_tabla_excel(self._pag_tabla)
        header = self._pag_tabla.horizontalHeader()
        header.setSortIndicatorShown(True)
        header.setSectionsClickable(True)
        header.sectionClicked.connect(self._on_header_clicked)
        self._pag_tabla.itemSelectionChanged.connect(self._sincronizar_seleccion)
        self._pag_tabla.cellDoubleClicked.connect(lambda _r, _c: self._emit_doble())
        self._stack.addWidget(self._pag_tabla)

        self._pag_iconos = QScrollArea()
        self._pag_iconos.setWidgetResizable(True)
        self._cont_iconos = QWidget()
        self._grid = QGridLayout(self._cont_iconos)
        self._grid.setContentsMargins(4, 4, 4, 4)
        self._grid.setSpacing(12)
        self._grid.setAlignment(Qt.AlignTop)
        self._pag_iconos.setWidget(self._cont_iconos)
        self._stack.addWidget(self._pag_iconos)

        lay.addWidget(self._stack, 1)
        self.set_vista("tabla")

    def _crear_boton(self, glifo: str, texto: str, color: str, fn) -> QPushButton:
        btn = QPushButton(texto)
        btn.setIcon(mono_icon(glifo, 16, color))
        btn.setCursor(Qt.PointingHandCursor)
        btn.clicked.connect(fn)
        return btn

    def _crear_boton_vista(self, glifo: str, texto: str) -> QPushButton:
        btn = QPushButton(texto)
        btn.setObjectName("viewSwitch")
        btn.setIcon(mono_icon(glifo, 18, "#475569"))
        btn.setCheckable(True)
        btn.setCursor(Qt.PointingHandCursor)
        btn.clicked.connect(lambda _c, v=texto.lower(): self.set_vista(v))
        return btn

    # ------------------------------------------------- Configuración desde código
    def set_columnas(self, columnas: list[dict]) -> None:
        self._col_config = list(columnas or [])
        self._cmb_agrupar.clear()
        self._cmb_agrupar.addItem("Sin agrupar", None)
        for c in self._col_config:
            self._cmb_agrupar.addItem(c.get("titulo", c.get("key", "")), c.get("key"))

    def set_datos(self, registros) -> None:
        self._registros = list(registros or [])
        self._aplicar_filtro()

    def set_renderers(self, fila=None, claves=None, tarjeta=None, lista=None) -> None:
        self._fila_fn = fila
        self._claves_fn = claves
        self._tarjeta_fn = tarjeta
        self._lista_fn = lista

    def set_acciones(self, acciones: list[dict]) -> None:
        self._acciones = list(acciones or [])

    def set_filtros(self, filtros: list) -> None:
        self._filtros = list(filtros or [])
        if self._registros:
            self._aplicar_filtro()

    def set_agrupacion(self, clave: str | None) -> None:
        self._agrupar_por = clave
        idx = self._cmb_agrupar.findData(clave)
        if idx >= 0:
            self._cmb_agrupar.setCurrentIndex(idx)
        self._render()

    def set_plantilla_excel(self, ruta: str | None, inicio: str = "A3") -> None:
        self._plantilla_excel = ruta
        self._plantilla_inicio = inicio

    def set_reporte_config(self, config: dict) -> None:
        self._reporte_cfg = dict(config or {})

    def set_buscador_visible(self, visible: bool) -> None:
        self._txt_buscar.setVisible(visible)

    def buscar(self, texto: str) -> None:
        self._txt_buscar.setText(texto)

    def registro_seleccionado(self):
        idx = self._stack.currentIndex()
        if idx == self._idx_iconos:
            return getattr(self, "_seleccionado", None)
        if idx == self._idx_tabla:
            fila = self._pag_tabla.currentRow()
            if 0 <= fila < len(self._mapa_rec):
                return self._mapa_rec[fila]
            return None
        fila = self._pag_lista.currentRow()
        if 0 <= fila < len(self._visibles):
            return self._visibles[fila]
        return None

    def datos_visibles(self) -> list:
        return list(self._visibles)

    def set_vista(self, vista: str) -> None:
        mapa = {
            "lista": (self._idx_lista, self._btn_lista),
            "tabla": (self._idx_tabla, self._btn_tabla),
            "iconos": (self._idx_iconos, self._btn_iconos),
        }
        if vista not in mapa:
            return
        idx, btn = mapa[vista]
        btn.setChecked(True)
        self._stack.setCurrentIndex(idx)
        if vista == "iconos":
            self._render_iconos()

    @property
    def table(self) -> QTableWidget:
        return self._pag_tabla

    # --------------------------------------------------------- Filtros y búsqueda
    def _on_buscar(self, texto: str) -> None:
        self._buscar = texto.strip().lower()
        self._aplicar_filtro()

    def _on_cmb_agrupar(self) -> None:
        self._agrupar_por = self._cmb_agrupar.currentData()
        self._render()

    def _pasa_filtros(self, rec) -> bool:
        for fn in self._filtros:
            if not fn(rec):
                return False
        if self._buscar:
            fila = self._fila_fn(rec) if self._fila_fn else self._fila_por_defecto(rec)
            hay = any(self._buscar in str(x).lower() for x in fila)
            if not hay:
                return False
        return True

    def _fila_por_defecto(self, rec) -> list:
        return [str(rec.get(c["key"], "")) for c in self._col_config]

    def _aplicar_filtro(self) -> None:
        self._visibles = [r for r in self._registros if self._pasa_filtros(r)]
        if self._sort_col >= 0:
            self._ordenar(self._sort_col, self._sort_asc,
                          actualizar_indicator=False)
        self._render()

    # --------------------------------------------------------------- Agrupación
    def _agrupar(self):
        if not self._agrupar_por:
            return None
        grupos: dict = {}
        for rec in self._visibles:
            valor = self._valor_columna(rec, self._agrupar_por)
            grupos.setdefault(valor, []).append(rec)
        orden = sorted(grupos.keys(), key=lambda v: str(v).lower())
        return [(v, grupos[v]) for v in orden]

    # -------------------------------------------------------------------- Render
    def _render(self) -> None:
        self._render_tabla()
        self._render_lista()
        self._render_iconos()
        n = len(self._visibles)
        info = f"{n} registro{'s' if n != 1 else ''}"
        if self._filtros:
            info += "  ·  filtros activos"
        self._lbl_estado.setText(info)

    def _render_tabla(self) -> None:
        t = self._pag_tabla
        n_cols = len(self._col_config) + (1 if self._acciones else 0)
        t.clear()
        t.setColumnCount(n_cols)
        t.setHorizontalHeaderLabels(
            [c.get("titulo", c.get("key", "")) for c in self._col_config]
            + ([""] if self._acciones else []))
        t.setRowCount(0)
        self._mapa_rec = []
        for i, c in enumerate(self._col_config):
            t.setColumnWidth(i, c.get("ancho", 110))
        if self._acciones:
            t.setColumnWidth(n_cols - 1, max(48, 44 * len(self._acciones) + 8))
            t.horizontalHeader().setSectionResizeMode(
                n_cols - 1, QHeaderView.Fixed)

        grupos = self._agrupar()
        for valor, recs in grupos or [(None, self._visibles)]:
            if valor is not None:
                self._insertar_fila_grupo(t, valor, n_cols)
            for rec in recs:
                self._insertar_fila(t, rec)

    def _insertar_fila_grupo(self, t: QTableWidget, valor, n_cols: int) -> None:
        r = t.rowCount()
        t.insertRow(r)
        item = _ItemOrdenable(f"{self._agrupar_por or ''}: {valor}")
        item.setData(Qt.UserRole, str(valor).lower())
        item.setFlags(Qt.ItemIsEnabled)
        item.setBackground(QColor("#eef2ff"))
        item.setForeground(QColor("#4f46e5"))
        fuente = QFont()
        fuente.setBold(True)
        item.setFont(fuente)
        t.setItem(r, 0, item)
        t.setSpan(r, 0, 1, n_cols)
        self._mapa_rec.append(None)

    def _insertar_fila(self, t: QTableWidget, rec) -> None:
        r = t.rowCount()
        t.insertRow(r)
        if self._acciones:
            t.setRowHeight(r, t.verticalHeader().defaultSectionSize() * 2)
        fila = self._fila_fn(rec) if self._fila_fn else self._fila_por_defecto(rec)
        claves = self._claves_fn(rec) if self._claves_fn else None
        for c, cfg in enumerate(self._col_config):
            texto = fila[c] if c < len(fila) else ""
            item = _ItemOrdenable(str(texto))
            if claves and c < len(claves):
                clave = claves[c]
            elif isinstance(texto, (int, float)):
                clave = float(texto)
            else:
                clave = str(texto).lower()
            item.setData(Qt.UserRole, clave)
            if cfg.get("tipo") == "numero":
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            t.setItem(r, c, item)
        if self._acciones:
            t.setCellWidget(r, len(self._col_config),
                            self._widget_acciones(rec))
        self._mapa_rec.append(rec)

    def _widget_acciones(self, rec) -> QWidget:
        cont = QWidget()
        lay = QHBoxLayout(cont)
        lay.setContentsMargins(2, 2, 2, 2)
        lay.setSpacing(4)
        for acc in self._acciones:
            btn = QToolButton()
            btn.setText(acc.get("texto", ""))
            btn.setIcon(mono_icon(acc.get("icono", "mas"), 14,
                                  acc.get("color", "#4f46e5")))
            btn.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(partial(self._ejecutar_accion, acc, rec))
            lay.addWidget(btn)
        return cont

    def _ejecutar_accion(self, acc: dict, rec) -> None:
        cb = acc.get("callback")
        if cb:
            cb(rec)

    def _render_lista(self) -> None:
        l = self._pag_lista
        l.blockSignals(True)
        l.clear()
        for rec in self._visibles:
            if self._lista_fn:
                titulo, subtitulo = self._lista_fn(rec)
            else:
                fila = self._fila_fn(rec) if self._fila_fn else self._fila_por_defecto(rec)
                titulo = str(fila[0]) if fila else ""
                subtitulo = "  |  ".join(str(x) for x in fila[1:])
            it = QListWidgetItem(f"{titulo}\n{subtitulo}")
            it.setToolTip(subtitulo)
            l.addItem(it)
        l.blockSignals(False)

    def _render_iconos(self) -> None:
        while self._grid.count():
            w = self._grid.takeAt(0).widget()
            if w:
                w.deleteLater()
        cols = max(1, self._cont_iconos.width() // 230)
        for i, rec in enumerate(self._visibles):
            card = _Tarjeta()
            datos = self._tarjeta_fn(rec) if self._tarjeta_fn else {}
            tile = datos.get("tile")
            if tile:
                try:
                    from src.utils.icons import tile_icon
                    pixmap = tile_icon(tile, 44).pixmap(44, 44)
                except KeyError:
                    pixmap = mono_icon(tile, 44,
                                       datos.get("color", "#4f46e5")).pixmap(44, 44)
                card._icono.setPixmap(pixmap)
            else:
                card._icono.setPixmap(mono_icon(
                    datos.get("icono", "oc"), 44,
                    datos.get("color", "#4f46e5")).pixmap(44, 44))
            card._titulo.setText(datos.get("titulo", ""))
            card._subtitulo.setText(datos.get("subtitulo", ""))
            badge = datos.get("badge", "")
            card._badge.setText(badge)
            card._badge.setVisible(bool(badge))
            card.setProperty("rec", rec)
            card.clicked.connect(lambda r=rec: self._seleccionar_tarjeta(r))
            card.doubleClicked.connect(self._emit_doble)
            self._grid.addWidget(card, i // cols, i % cols)

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        if self._stack.currentIndex() == self._idx_iconos:
            self._render_iconos()

    # ---------------------------------------------------------------- Ordenación
    def _on_header_clicked(self, col: int) -> None:
        if col >= len(self._col_config):
            return
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self._ordenar(self._sort_col, self._sort_asc)

    def _ordenar(self, col: int, asc: bool, actualizar_indicator: bool = True) -> None:
        def clave(rec):
            claves = self._claves_fn(rec) if self._claves_fn else None
            if claves and col < len(claves):
                k = claves[col]
            else:
                fila = self._fila_fn(rec) if self._fila_fn else self._fila_por_defecto(rec)
                k = fila[col] if col < len(fila) else ""
            if k is None:
                return (1, "")
            return (0, k)

        self._visibles.sort(key=clave, reverse=not asc)
        if actualizar_indicator:
            self._pag_tabla.horizontalHeader().setSortIndicator(
                col, Qt.AscendingOrder if asc else Qt.DescendingOrder)
        self._render()

    def _valor_columna(self, rec, key: str):
        keys = [c.get("key") for c in self._col_config]
        if key in keys:
            fila = self._fila_fn(rec) if self._fila_fn else self._fila_por_defecto(rec)
            i = keys.index(key)
            return fila[i] if i < len(fila) else ""
        return rec.get(key, "") if isinstance(rec, dict) else ""

    # --------------------------------------------------------------- Selección
    def _seleccionar_tarjeta(self, rec) -> None:
        self._seleccionado = rec
        for i in range(self._grid.count()):
            w = self._grid.itemAt(i).widget()
            if isinstance(w, _Tarjeta):
                sel = w.property("rec") is rec
                w.setProperty("selected", sel)
                w.style().unpolish(w)
                w.style().polish(w)
        self.selectionChanged.emit()

    def _sincronizar_seleccion(self) -> None:
        self.selectionChanged.emit()

    def _emit_doble(self) -> None:
        if self.registro_seleccionado() is not None:
            self.doubleClicked.emit()

    # ------------------------------------------------------ Exportación e impresión
    def exportar_excel(self) -> None:
        self._render_tabla()
        if self._plantilla_excel:
            self._exportar_excel_plantilla()
            return
        titulo = self._reporte_cfg.get("titulo", "Reporte")
        export_table_to_excel(self._pag_tabla, titulo, self)

    def _exportar_excel_plantilla(self) -> None:
        titulo = self._reporte_cfg.get("titulo", "Reporte")
        ruta, _ = QFileDialog.getSaveFileName(
            self, "Exportar (plantilla)", f"{titulo}.xlsx", "Excel (*.xlsx)")
        if not ruta:
            return
        wb = load_workbook(self._plantilla_excel)
        ws = wb.active
        fila, col = _celda_inicio(self._plantilla_inicio)
        for j, cfg in enumerate(self._col_config):
            ws.cell(row=fila, column=col + j,
                    value=cfg.get("titulo", cfg.get("key", "")))
        for i, rec in enumerate(self._visibles):
            valores = self._fila_fn(rec) if self._fila_fn else self._fila_por_defecto(rec)
            for j in range(len(self._col_config)):
                ws.cell(row=fila + 1 + i, column=col + j,
                        value=valores[j] if j < len(valores) else "")
        wb.save(ruta)

    def exportar_pdf(self) -> None:
        self._render_tabla()
        titulo = self._reporte_cfg.get("titulo", "Reporte")
        print_table(self._pag_tabla, titulo, self)

    def imprimir(self) -> None:
        if self._reporte_cfg.get("titulo"):
            self._imprimir_reporte_personalizado()
        else:
            self.exportar_pdf()

    def _imprimir_reporte_personalizado(self) -> None:
        titulo = self._reporte_cfg.get("titulo", "Reporte")
        ruta, _ = QFileDialog.getSaveFileName(
            self, f"Imprimir - {titulo}", f"{titulo}.pdf", "PDF (*.pdf)")
        if not ruta:
            return
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageSize(QPrinter.Letter)
        printer.setOrientation(QPrinter.Landscape)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(ruta)
        doc = QTextDocument()
        doc.setHtml(self._reporte_html())
        doc.print_(printer)

    def _reporte_html(self) -> str:
        titulo = self._reporte_cfg.get("titulo", "Reporte")
        subtitulo = self._reporte_cfg.get("subtitulo", "")
        cols = [c.get("titulo", c.get("key", "")) for c in self._col_config]
        cuerpo = ""
        grupos = self._agrupar()
        for valor, recs in grupos or [(None, self._visibles)]:
            if valor is not None:
                cuerpo += (f'<tr><td colspan="{len(cols)}" '
                           f'style="background:#eef2ff;color:#4f46e5;'
                           f'font-weight:bold;">{_esc(str(valor))}</td></tr>')
            for rec in recs:
                fila = self._fila_fn(rec) if self._fila_fn else self._fila_por_defecto(rec)
                celdas = "".join(f"<td>{_esc(str(x))}</td>" for x in fila)
                cuerpo += f"<tr>{celdas}</tr>"
        return (
            "<html><head><meta charset='utf-8'></head><body>"
            f"<h1 style='margin:0'>{_esc(str(titulo))}</h1>"
            f"<p style='color:#64748b;margin:2px 0 12px'>{_esc(str(subtitulo))}</p>"
            f"<table border='1' cellspacing='0' cellpadding='5' "
            f"style='border-collapse:collapse;width:100%;font-size:11px'>"
            f"<tr>{''.join(f'<th style='
            f'\"background:#111827;color:#fff;text-align:left\">{_esc(c)}</th>' for c in cols)}</tr>"
            f"{cuerpo}</table></body></html>"
        )
